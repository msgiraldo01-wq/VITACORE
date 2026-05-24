# =============================================
# VITACORE · Blueprint Cartera
# Archivo: blueprints/bp_financiero/cartera/cartera.py
# =============================================

from flask import Blueprint, render_template, request, jsonify, session, Response
from repositories.fin_cartera_repo import (
    obtener_todas_facturas,
    obtener_factura_por_numero,
    obtener_pagos_por_factura,
    registrar_pago,
    obtener_kpis_cartera,
    actualizar_dias_mora,
    obtener_facturas_para_excel,
)
from repositories.fin_glosas_repo import obtener_glosas_por_factura
from repositories.fin_cartera_documentos_repo import (
    obtener_documentos_por_factura,
    obtener_todos_documentos,
    crear_documento,
    anular_documento,
    obtener_kpis_documentos,
    ETIQUETAS,
)
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import date

bp_financiero_cartera = Blueprint(
    "cartera", __name__, url_prefix="/financiero/cartera"
)


# --------------------------------------------------
# LISTA DE CARTERA
# --------------------------------------------------
@bp_financiero_cartera.route("/")
def index():
    actualizar_dias_mora()
    facturas = obtener_todas_facturas()
    kpis     = obtener_kpis_cartera()
    return render_template(
        "financiero/cartera/cartera.html",
        facturas=facturas,
        kpis=kpis,
    )


# --------------------------------------------------
# DETALLE DE FACTURA
# --------------------------------------------------
@bp_financiero_cartera.route("/detalle/<numero_factura>")
def detalle(numero_factura):
    factura    = obtener_factura_por_numero(numero_factura)
    pagos      = obtener_pagos_por_factura(numero_factura)
    glosas     = obtener_glosas_por_factura(numero_factura)
    documentos = obtener_documentos_por_factura(numero_factura)

    if not factura:
        return "Factura no encontrada", 404

    return render_template(
        "financiero/cartera/cartera_detalle.html",
        data=factura,
        pagos=pagos,
        glosas=glosas,
        documentos=documentos,
        etiquetas=ETIQUETAS,
    )


# --------------------------------------------------
# MÓDULO DE DOCUMENTOS (lista general)
# --------------------------------------------------
@bp_financiero_cartera.route("/documentos")
def documentos():
    docs = obtener_todos_documentos()
    kpis = obtener_kpis_documentos()
    return render_template(
        "financiero/cartera/documentos.html",
        documentos=docs,
        kpis=kpis,
        etiquetas=ETIQUETAS,
    )


# --------------------------------------------------
# REGISTRAR PAGO (RC)
# --------------------------------------------------
@bp_financiero_cartera.route("/registrar-pago", methods=["POST"])
def api_registrar_pago():
    factura_id     = request.form.get("factura_id")
    numero_factura = request.form.get("numero_factura")
    valor_pago     = request.form.get("valor_pago")
    fecha_pago     = request.form.get("fecha_pago")

    if not all([factura_id, numero_factura, valor_pago, fecha_pago]):
        return jsonify({"ok": False, "error": "Faltan campos obligatorios"}), 400

    usuario = session.get("user", {}).get("username", "Sistema")

    data = {
        "factura_id":          factura_id,
        "numero_factura":      numero_factura,
        "valor_pago":          valor_pago,
        "fecha_pago":          fecha_pago,
        "banco":               request.form.get("banco"),
        "tipo_pago":           request.form.get("tipo_pago", "Transferencia"),
        "referencia_bancaria": request.form.get("referencia_bancaria"),
        "observacion":         request.form.get("observacion"),
        "registrado_por":      usuario,
    }

    archivo = request.files.get("soporte")

    try:
        registrar_pago(data, archivo=archivo)
        return jsonify({"ok": True, "mensaje": "Pago registrado correctamente"})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


# --------------------------------------------------
# REGISTRAR DOCUMENTO (NC, ND, DEV, ACL, CRU)
# --------------------------------------------------
@bp_financiero_cartera.route("/registrar-documento", methods=["POST"])
def api_registrar_documento():
    factura_id     = request.form.get("factura_id")
    numero_factura = request.form.get("numero_factura")
    tipo_documento = request.form.get("tipo_documento")
    valor          = request.form.get("valor", 0)
    fecha          = request.form.get("fecha_documento")

    if not all([factura_id, numero_factura, tipo_documento, fecha]):
        return jsonify({"ok": False, "error": "Faltan campos obligatorios"}), 400

    usuario = session.get("user", {}).get("username", "Sistema")

    data = {
        "factura_id":         factura_id,
        "numero_factura":     numero_factura,
        "tipo_documento":     tipo_documento,
        "eps":                request.form.get("eps"),
        "valor":              float(valor) if valor else 0,
        "fecha_documento":    fecha,
        "descripcion":        request.form.get("descripcion"),
        "referencia_externa": request.form.get("referencia_externa"),
        "afecta_saldo":       request.form.get("afecta_saldo", "true") == "true",
        "registrado_por":     usuario,
    }

    archivo = request.files.get("soporte")

    try:
        numero = crear_documento(data, archivo=archivo)
        return jsonify({"ok": True, "numero_documento": numero})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


# --------------------------------------------------
# ANULAR DOCUMENTO
# --------------------------------------------------
@bp_financiero_cartera.route("/anular-documento", methods=["POST"])
def api_anular_documento():
    body             = request.get_json(silent=True) or {}
    numero_documento = body.get("numero_documento")

    if not numero_documento:
        return jsonify({"ok": False, "error": "Falta número de documento"}), 400

    usuario = session.get("user", {}).get("username", "Sistema")

    try:
        anular_documento(numero_documento, usuario)
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


# --------------------------------------------------
# EXPORTAR EXCEL CARTERA
# --------------------------------------------------
@bp_financiero_cartera.route("/exportar-excel")
def exportar_excel():
    facturas = obtener_facturas_para_excel()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cartera"

    color_header   = "0D5C63"
    fuente_header  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    fuente_normal  = Font(name="Calibri", size=10)
    relleno_header = PatternFill("solid", fgColor=color_header)
    relleno_alt    = PatternFill("solid", fgColor="F0F7F8")
    borde = Border(
        left=Side(style="thin",   color="D9E4E7"),
        right=Side(style="thin",  color="D9E4E7"),
        top=Side(style="thin",    color="D9E4E7"),
        bottom=Side(style="thin", color="D9E4E7"),
    )
    centro  = Alignment(horizontal="center", vertical="center")
    derecha = Alignment(horizontal="right",  vertical="center")

    ws.merge_cells("A1:M1")
    ws["A1"].value     = "VITACORE · REPORTE DE CARTERA HOSPITALARIA"
    ws["A1"].font      = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    ws["A1"].fill      = PatternFill("solid", fgColor=color_header)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:M2")
    ws["A2"].value     = f"Generado: {date.today().strftime('%d/%m/%Y')}  ·  Total: {len(facturas)}"
    ws["A2"].font      = Font(name="Calibri", size=10, color="5E7278")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    encabezados = [
        "Factura", "EPS", "NIT EPS",
        "Valor factura", "Pagado", "Glosas",
        "Saldo pendiente", "Días mora", "Estado",
        "F. radicación", "F. vencimiento", "Último pago", "Responsable"
    ]
    anchos = [18, 20, 16, 16, 14, 12, 16, 10, 14, 14, 14, 14, 18]

    for col, (titulo, ancho) in enumerate(zip(encabezados, anchos), start=1):
        celda = ws.cell(row=3, column=col, value=titulo)
        celda.font      = fuente_header
        celda.fill      = relleno_header
        celda.alignment = centro
        celda.border    = borde
        ws.column_dimensions[celda.column_letter].width = ancho

    ws.row_dimensions[3].height = 24

    total_factura = total_pagado = total_glosas = total_saldo = 0

    for idx, f in enumerate(facturas, start=4):
        relleno_fila = relleno_alt if idx % 2 == 0 else PatternFill()
        valores = [
            f.get("numero_factura", ""),
            f.get("eps", ""),
            f.get("nit_eps", ""),
            f.get("valor_factura", 0),
            f.get("valor_pagado", 0),
            f.get("valor_glosas", 0),
            f.get("saldo_pendiente", 0),
            f.get("dias_mora", 0),
            f.get("estado", "").replace("_", " ").title(),
            f.get("fecha_radicacion", ""),
            f.get("fecha_vencimiento", ""),
            f.get("ultimo_pago", ""),
            f.get("responsable", ""),
        ]
        for col, valor in enumerate(valores, start=1):
            celda = ws.cell(row=idx, column=col, value=valor)
            celda.font   = fuente_normal
            celda.fill   = relleno_fila
            celda.border = borde
            if col in [4, 5, 6, 7]:
                celda.number_format = '$#,##0'
                celda.alignment = derecha
            elif col == 8:
                celda.alignment = centro
            else:
                celda.alignment = Alignment(vertical="center")

        total_factura += f.get("valor_factura",   0) or 0
        total_pagado  += f.get("valor_pagado",    0) or 0
        total_glosas  += f.get("valor_glosas",    0) or 0
        total_saldo   += f.get("saldo_pendiente", 0) or 0

    fila_total    = len(facturas) + 4
    relleno_total = PatternFill("solid", fgColor="0D5C63")

    ws.merge_cells(f"A{fila_total}:C{fila_total}")
    ws[f"A{fila_total}"].value     = "TOTALES"
    ws[f"A{fila_total}"].font      = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    ws[f"A{fila_total}"].fill      = relleno_total
    ws[f"A{fila_total}"].alignment = centro
    ws[f"A{fila_total}"].border    = borde

    for col, total in enumerate([total_factura, total_pagado, total_glosas, total_saldo], start=4):
        celda = ws.cell(row=fila_total, column=col, value=total)
        celda.font          = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        celda.fill          = relleno_total
        celda.number_format = '$#,##0'
        celda.alignment     = derecha
        celda.border        = borde

    for col in range(8, 14):
        celda = ws.cell(row=fila_total, column=col, value="")
        celda.fill   = relleno_total
        celda.border = borde

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:M{fila_total - 1}"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    nombre = f"cartera_vitacore_{date.today().strftime('%Y%m%d')}.xlsx"
    return Response(
        buffer.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={nombre}"}
    )