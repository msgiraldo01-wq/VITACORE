# =============================================
# VITACORE · Blueprint Glosas
# Archivo: blueprints/bp_financiero/glosas/glosas.py
# =============================================

from flask import Blueprint, render_template, request, jsonify, session, Response
from repositories.fin_glosas_repo import (
    obtener_todas_glosas,
    obtener_glosa_por_numero,
    obtener_respuestas_por_glosa,
    registrar_respuesta,
    obtener_kpis_glosas,
    obtener_glosas_para_excel,
)
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import date

bp_financiero_glosas = Blueprint(
    "bp_financiero_glosas",
    __name__,
    url_prefix="/financiero/glosas"
)


# --------------------------------------------------
# LISTA DE GLOSAS
# --------------------------------------------------
@bp_financiero_glosas.route("/")
def glosas():
    kpis   = obtener_kpis_glosas()
    glosas = obtener_todas_glosas()
    return render_template(
        "financiero/glosas/glosas.html",
        kpis=kpis,
        glosas=glosas,
    )


# --------------------------------------------------
# DETALLE DE GLOSA
# --------------------------------------------------
@bp_financiero_glosas.route("/detalle/<numero_glosa>")
def detalle(numero_glosa):
    glosa     = obtener_glosa_por_numero(numero_glosa)
    respuestas = obtener_respuestas_por_glosa(numero_glosa)

    if not glosa:
        return "Glosa no encontrada", 404

    return render_template(
        "financiero/glosas/glosa_detalle.html",
        glosa=glosa,
        respuestas=respuestas,
    )


# --------------------------------------------------
# REGISTRAR RESPUESTA (API)
# --------------------------------------------------
@bp_financiero_glosas.route("/registrar-respuesta", methods=["POST"])
def api_registrar_respuesta():
    glosa_id      = request.form.get("glosa_id")
    numero_glosa  = request.form.get("numero_glosa")
    descripcion   = request.form.get("descripcion")

    if not all([glosa_id, numero_glosa, descripcion]):
        return jsonify({"ok": False, "error": "Faltan campos obligatorios"}), 400

    usuario = session.get("user", {}).get("username", "Sistema")

    data = {
        "glosa_id":        glosa_id,
        "numero_glosa":    numero_glosa,
        "tipo_respuesta":  request.form.get("tipo_respuesta", "respuesta"),
        "descripcion":     descripcion,
        "valor_propuesto": request.form.get("valor_propuesto", 0),
        "nuevo_estado":    request.form.get("nuevo_estado", "en_respuesta"),
        "registrado_por":  usuario,
    }

    archivo = request.files.get("soporte")

    try:
        registrar_respuesta(data, archivo=archivo)
        return jsonify({"ok": True, "mensaje": "Respuesta registrada correctamente"})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


# --------------------------------------------------
# EXPORTAR EXCEL
# --------------------------------------------------
@bp_financiero_glosas.route("/exportar-excel")
def exportar_excel():
    glosas = obtener_glosas_para_excel()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Glosas"

    color_header = "0D5C63"
    fuente_header = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    fuente_normal = Font(name="Calibri", size=10)
    relleno_header = PatternFill("solid", fgColor=color_header)
    relleno_alt    = PatternFill("solid", fgColor="F0F7F8")
    borde = Border(
        left=Side(style="thin", color="D9E4E7"),
        right=Side(style="thin", color="D9E4E7"),
        top=Side(style="thin", color="D9E4E7"),
        bottom=Side(style="thin", color="D9E4E7"),
    )
    centro  = Alignment(horizontal="center", vertical="center")
    derecha = Alignment(horizontal="right",  vertical="center")

    # Título
    ws.merge_cells("A1:O1")
    ws["A1"].value     = "VITACORE · REPORTE DE GLOSAS HOSPITALARIAS"
    ws["A1"].font      = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    ws["A1"].fill      = PatternFill("solid", fgColor=color_header)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:O2")
    ws["A2"].value     = f"Generado: {date.today().strftime('%d/%m/%Y')}  ·  Total glosas: {len(glosas)}"
    ws["A2"].font      = Font(name="Calibri", size=10, color="5E7278")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    encabezados = [
        "Glosa", "Factura", "EPS", "Tipo",
        "Valor glosado", "Aceptado", "Levantado", "Ratificado",
        "F. Glosa", "F. Vencimiento", "F. Respuesta",
        "Días", "Estado", "Causal", "Respondida por"
    ]
    anchos = [16, 18, 20, 20, 14, 12, 12, 12, 14, 14, 14, 8, 14, 30, 20]

    for col, (titulo, ancho) in enumerate(zip(encabezados, anchos), start=1):
        celda = ws.cell(row=3, column=col, value=titulo)
        celda.font      = fuente_header
        celda.fill      = relleno_header
        celda.alignment = centro
        celda.border    = borde
        ws.column_dimensions[celda.column_letter].width = ancho

    ws.row_dimensions[3].height = 24

    total_glosado = total_aceptado = total_levantado = total_ratificado = 0

    for idx, g in enumerate(glosas, start=4):
        relleno_fila = relleno_alt if idx % 2 == 0 else PatternFill()
        valores = [
            g.get("numero_glosa", ""),
            g.get("numero_factura", ""),
            g.get("eps", ""),
            g.get("tipo_glosa", ""),
            g.get("valor_glosado", 0),
            g.get("valor_aceptado", 0),
            g.get("valor_levantado", 0),
            g.get("valor_ratificado", 0),
            g.get("fecha_glosa", ""),
            g.get("fecha_vencimiento", ""),
            g.get("fecha_respuesta", ""),
            g.get("dias_respuesta", 0),
            g.get("estado", "").replace("_", " ").title(),
            g.get("causal", ""),
            g.get("respondida_por", ""),
        ]
        for col, valor in enumerate(valores, start=1):
            celda = ws.cell(row=idx, column=col, value=valor)
            celda.font   = fuente_normal
            celda.fill   = relleno_fila
            celda.border = borde
            if col in [5, 6, 7, 8]:
                celda.number_format = '$#,##0'
                celda.alignment = derecha
            elif col == 12:
                celda.alignment = centro
            else:
                celda.alignment = Alignment(vertical="center")

        total_glosado   += g.get("valor_glosado",   0) or 0
        total_aceptado  += g.get("valor_aceptado",  0) or 0
        total_levantado += g.get("valor_levantado", 0) or 0
        total_ratificado += g.get("valor_ratificado", 0) or 0

    # Totales
    fila_total    = len(glosas) + 4
    relleno_total = PatternFill("solid", fgColor="0D5C63")

    ws.merge_cells(f"A{fila_total}:D{fila_total}")
    ws[f"A{fila_total}"].value     = "TOTALES"
    ws[f"A{fila_total}"].font      = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    ws[f"A{fila_total}"].fill      = relleno_total
    ws[f"A{fila_total}"].alignment = centro
    ws[f"A{fila_total}"].border    = borde

    for col, total in enumerate([total_glosado, total_aceptado, total_levantado, total_ratificado], start=5):
        celda = ws.cell(row=fila_total, column=col, value=total)
        celda.font          = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        celda.fill          = relleno_total
        celda.number_format = '$#,##0'
        celda.alignment     = derecha
        celda.border        = borde

    for col in range(9, 16):
        celda = ws.cell(row=fila_total, column=col, value="")
        celda.fill   = relleno_total
        celda.border = borde

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:O{fila_total - 1}"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    nombre = f"glosas_vitacore_{date.today().strftime('%Y%m%d')}.xlsx"
    return Response(
        buffer.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={nombre}"}
    )
    
@bp_financiero_glosas.route("/nueva", methods=["POST"])
def api_nueva_glosa():
    numero_factura = request.form.get("numero_factura")
    eps            = request.form.get("eps")
    tipo_glosa     = request.form.get("tipo_glosa")
    valor_glosado  = request.form.get("valor_glosado")
    fecha_glosa    = request.form.get("fecha_glosa")
    causal         = request.form.get("causal")
    observaciones  = request.form.get("observaciones")

    if not all([numero_factura, eps, tipo_glosa, valor_glosado, fecha_glosa]):
        return jsonify({"ok": False, "error": "Faltan campos obligatorios"}), 400

    try:
        from repositories.fin_glosas_repo import crear_glosa
        resultado = crear_glosa({
            "numero_factura": numero_factura,
            "eps":            eps,
            "tipo_glosa":     tipo_glosa,
            "valor_glosado":  float(valor_glosado),
            "fecha_glosa":    fecha_glosa,
            "causal":         causal,
            "observaciones":  observaciones,
            "registrado_por": session.get("user", {}).get("username", "Sistema"),
        })
        return jsonify({"ok": True, "numero_glosa": resultado})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500