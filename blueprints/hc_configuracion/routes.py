from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, Response
from repositories import hc_sedes_repo as repo
from repositories import hc_tipos_documento_repo as repo_td
from repositories import hc_especialidades_repo as repo_esp
from repositories import hc_consultorios_repo as repo_cons
from repositories import hc_sedes_repo as repo_sedes
from repositories import hc_profesionales_repo as repo_prof
from repositories import hc_servicios_repo
from repositories import hc_paises_repo as repo_paises
from repositories import hc_cie10_repo as repo_cie10
from repositories import hc_departamentos_repo as repo_dep
from repositories import hc_municipios_repo as repo_muni
from repositories import hc_eps_repo as repo_eps
from repositories import hc_recursos_repo as repo_recursos
from repositories import hc_agendas_repo as agendas_repo
from repositories import hc_cups_repo as repo_cups
import repositories.prof_procedimientos_repository as prof_proc_repo
import repositories.rec_procedimientos_repository  as rec_proc_repo
import repositories.hc_medicos_remitentes_repo as repo_mr
import repositories.hc_prof_programacion_repo as repo_prog
import io
import csv
import repositories.hc_clientes_repo as repo_clientes
import repositories.hc_contratos_repo as repo_contratos
import repositories.hc_manuales_repo as repo_manuales
from services.supabase_service import get_supabase_public


bp_hc_configuracion = Blueprint(
    "hc_configuracion",
    __name__,
    url_prefix="/hc/configuracion"
)


@bp_hc_configuracion.route("/")
def index():
    return render_template("hc/configuracion/index.html")


# ══════════════════════════════════════════════════════════════════
#  SEDES
# ══════════════════════════════════════════════════════════════════

@bp_hc_configuracion.route("/sedes")
def sedes():
    items = repo.listar()
    return render_template("hc/configuracion/sedes.html", items=items)


@bp_hc_configuracion.route("/sedes/nueva", methods=["GET", "POST"])
def sede_nueva():
    if request.method == "POST":
        data = {
            "codigo": request.form.get("codigo"),
            "nombre": request.form.get("nombre"),
            "ciudad": request.form.get("ciudad"),
            "direccion": request.form.get("direccion"),
            "telefono": request.form.get("telefono"),
            "estado": request.form.get("estado"),
        }
        if not (data["codigo"] or "").strip():
            flash("El código es obligatorio.", "warning")
            return render_template("hc/configuracion/sedes_form.html", modo="crear", item=data)
        if not (data["nombre"] or "").strip():
            flash("El nombre es obligatorio.", "warning")
            return render_template("hc/configuracion/sedes_form.html", modo="crear", item=data)
        if repo.existe_codigo(data["codigo"]):
            flash("Ya existe una sede con ese código.", "warning")
            return render_template("hc/configuracion/sedes_form.html", modo="crear", item=data)
        repo.crear(data)
        flash("Sede creada correctamente.", "success")
        return redirect(url_for("hc_configuracion.sedes"))
    return render_template("hc/configuracion/sedes_form.html", modo="crear", item={})


@bp_hc_configuracion.route("/sedes/editar/<int:sede_id>", methods=["GET", "POST"])
def sede_editar(sede_id):
    item = repo.obtener(sede_id)
    if not item:
        flash("La sede no existe.", "error")
        return redirect(url_for("hc_configuracion.sedes"))
    if request.method == "POST":
        data = {
            "codigo": request.form.get("codigo"),
            "nombre": request.form.get("nombre"),
            "ciudad": request.form.get("ciudad"),
            "direccion": request.form.get("direccion"),
            "telefono": request.form.get("telefono"),
            "estado": request.form.get("estado"),
        }
        if not (data["codigo"] or "").strip():
            flash("El código es obligatorio.", "warning")
            return render_template("hc/configuracion/sedes_form.html", modo="editar", item={**item, **data})
        if not (data["nombre"] or "").strip():
            flash("El nombre es obligatorio.", "warning")
            return render_template("hc/configuracion/sedes_form.html", modo="editar", item={**item, **data})
        if repo.existe_codigo(data["codigo"], exclude_id=sede_id):
            flash("Ya existe otra sede con ese código.", "warning")
            return render_template("hc/configuracion/sedes_form.html", modo="editar", item={**item, **data})
        repo.actualizar(sede_id, data)
        flash("Sede actualizada correctamente.", "success")
        return redirect(url_for("hc_configuracion.sedes"))
    return render_template("hc/configuracion/sedes_form.html", modo="editar", item=item)


@bp_hc_configuracion.route("/sedes/toggle/<int:sede_id>", methods=["POST"])
def sede_toggle(sede_id):
    item = repo.obtener(sede_id)
    if not item:
        flash("La sede no existe.", "error")
        return redirect(url_for("hc_configuracion.sedes"))
    nuevo_estado = "INACTIVA" if item.get("estado") == "ACTIVA" else "ACTIVA"
    repo.cambiar_estado(sede_id, nuevo_estado)
    flash(f"Sede {'inactivada' if nuevo_estado == 'INACTIVA' else 'activada'} correctamente.", "success")
    return redirect(url_for("hc_configuracion.sedes"))


# ══════════════════════════════════════════════════════════════════
#  TIPOS DOCUMENTO
# ══════════════════════════════════════════════════════════════════

@bp_hc_configuracion.route("/tipos-documento")
def tipos_documento():
    items = repo_td.listar()
    return render_template("hc/configuracion/tipos_documento.html", items=items)


@bp_hc_configuracion.route("/tipos-documento/nuevo", methods=["GET", "POST"])
def tipo_documento_nuevo():
    if request.method == "POST":
        data = {
            "codigo": request.form.get("codigo"),
            "nombre": request.form.get("nombre"),
            "descripcion": request.form.get("descripcion"),
            "estado": request.form.get("estado"),
        }
        if not (data["codigo"] or "").strip():
            flash("El código es obligatorio.", "warning")
            return render_template("hc/configuracion/tipos_documento_form.html", modo="crear", item=data)
        if not (data["nombre"] or "").strip():
            flash("El nombre es obligatorio.", "warning")
            return render_template("hc/configuracion/tipos_documento_form.html", modo="crear", item=data)
        if repo_td.existe_codigo(data["codigo"]):
            flash("Ya existe un tipo de documento con ese código.", "warning")
            return render_template("hc/configuracion/tipos_documento_form.html", modo="crear", item=data)
        repo_td.crear(data)
        flash("Tipo de documento creado correctamente.", "success")
        return redirect(url_for("hc_configuracion.tipos_documento"))
    return render_template("hc/configuracion/tipos_documento_form.html", modo="crear", item={})


@bp_hc_configuracion.route("/tipos-documento/editar/<int:item_id>", methods=["GET", "POST"])
def tipo_documento_editar(item_id):
    item = repo_td.obtener(item_id)
    if not item:
        flash("El tipo de documento no existe.", "error")
        return redirect(url_for("hc_configuracion.tipos_documento"))
    if request.method == "POST":
        data = {
            "codigo": request.form.get("codigo"),
            "nombre": request.form.get("nombre"),
            "descripcion": request.form.get("descripcion"),
            "estado": request.form.get("estado"),
        }
        if not (data["codigo"] or "").strip():
            flash("El código es obligatorio.", "warning")
            return render_template("hc/configuracion/tipos_documento_form.html", modo="editar", item={**item, **data})
        if not (data["nombre"] or "").strip():
            flash("El nombre es obligatorio.", "warning")
            return render_template("hc/configuracion/tipos_documento_form.html", modo="editar", item={**item, **data})
        if repo_td.existe_codigo(data["codigo"], exclude_id=item_id):
            flash("Ya existe otro tipo de documento con ese código.", "warning")
            return render_template("hc/configuracion/tipos_documento_form.html", modo="editar", item={**item, **data})
        repo_td.actualizar(item_id, data)
        flash("Tipo de documento actualizado correctamente.", "success")
        return redirect(url_for("hc_configuracion.tipos_documento"))
    return render_template("hc/configuracion/tipos_documento_form.html", modo="editar", item=item)


@bp_hc_configuracion.route("/tipos-documento/toggle/<int:item_id>", methods=["POST"])
def tipo_documento_toggle(item_id):
    item = repo_td.obtener(item_id)
    if not item:
        flash("El tipo de documento no existe.", "error")
        return redirect(url_for("hc_configuracion.tipos_documento"))
    nuevo_estado = "INACTIVO" if item.get("estado") == "ACTIVO" else "ACTIVO"
    repo_td.cambiar_estado(item_id, nuevo_estado)
    flash(f"Tipo de documento {'inactivado' if nuevo_estado == 'INACTIVO' else 'activado'} correctamente.", "success")
    return redirect(url_for("hc_configuracion.tipos_documento"))


# ══════════════════════════════════════════════════════════════════
#  ESPECIALIDADES
# ══════════════════════════════════════════════════════════════════

@bp_hc_configuracion.route("/especialidades")
def especialidades():
    items = repo_esp.listar()
    return render_template("hc/configuracion/especialidades.html", items=items)


@bp_hc_configuracion.route("/especialidades/nueva", methods=["GET", "POST"])
def especialidad_nueva():
    if request.method == "POST":
        data = {
            "codigo": request.form.get("codigo"),
            "nombre": request.form.get("nombre"),
            "descripcion": request.form.get("descripcion"),
            "estado": request.form.get("estado"),
        }
        if not (data["codigo"] or "").strip():
            flash("El código es obligatorio.", "warning")
            return render_template("hc/configuracion/especialidades_form.html", modo="crear", item=data)
        if not (data["nombre"] or "").strip():
            flash("El nombre es obligatorio.", "warning")
            return render_template("hc/configuracion/especialidades_form.html", modo="crear", item=data)
        if repo_esp.existe_codigo(data["codigo"]):
            flash("Ya existe una especialidad con ese código.", "warning")
            return render_template("hc/configuracion/especialidades_form.html", modo="crear", item=data)
        repo_esp.crear(data)
        flash("Especialidad creada correctamente.", "success")
        return redirect(url_for("hc_configuracion.especialidades"))
    return render_template("hc/configuracion/especialidades_form.html", modo="crear", item={})


@bp_hc_configuracion.route("/especialidades/editar/<int:item_id>", methods=["GET", "POST"])
def especialidad_editar(item_id):
    item = repo_esp.obtener(item_id)
    if not item:
        flash("La especialidad no existe.", "error")
        return redirect(url_for("hc_configuracion.especialidades"))
    if request.method == "POST":
        data = {
            "codigo": request.form.get("codigo"),
            "nombre": request.form.get("nombre"),
            "descripcion": request.form.get("descripcion"),
            "estado": request.form.get("estado"),
        }
        if not (data["codigo"] or "").strip():
            flash("El código es obligatorio.", "warning")
            return render_template("hc/configuracion/especialidades_form.html", modo="editar", item={**item, **data})
        if not (data["nombre"] or "").strip():
            flash("El nombre es obligatorio.", "warning")
            return render_template("hc/configuracion/especialidades_form.html", modo="editar", item={**item, **data})
        if repo_esp.existe_codigo(data["codigo"], exclude_id=item_id):
            flash("Ya existe otra especialidad con ese código.", "warning")
            return render_template("hc/configuracion/especialidades_form.html", modo="editar", item={**item, **data})
        repo_esp.actualizar(item_id, data)
        flash("Especialidad actualizada correctamente.", "success")
        return redirect(url_for("hc_configuracion.especialidades"))
    return render_template("hc/configuracion/especialidades_form.html", modo="editar", item=item)


@bp_hc_configuracion.route("/especialidades/toggle/<int:item_id>", methods=["POST"])
def especialidad_toggle(item_id):
    item = repo_esp.obtener(item_id)
    if not item:
        flash("La especialidad no existe.", "error")
        return redirect(url_for("hc_configuracion.especialidades"))
    nuevo_estado = "INACTIVA" if item.get("estado") == "ACTIVA" else "ACTIVA"
    repo_esp.cambiar_estado(item_id, nuevo_estado)
    flash(f"Especialidad {'inactivada' if nuevo_estado == 'INACTIVA' else 'activada'} correctamente.", "success")
    return redirect(url_for("hc_configuracion.especialidades"))


# ══════════════════════════════════════════════════════════════════
#  CONSULTORIOS
# ══════════════════════════════════════════════════════════════════

@bp_hc_configuracion.route("/consultorios")
def consultorios():
    items = repo_cons.listar()
    return render_template("hc/configuracion/consultorios.html", items=items)


@bp_hc_configuracion.route("/consultorios/nuevo", methods=["GET", "POST"])
def consultorio_nuevo():
    sedes = repo_sedes.listar()
    if request.method == "POST":
        sede_id_raw = (request.form.get("sede_id") or "").strip()
        sede_id = int(sede_id_raw) if sede_id_raw.isdigit() else None
        sede_item = repo_sedes.obtener(sede_id) if sede_id else None
        data = {
            "sede_id": sede_id,
            "sede_nombre": (sede_item.get("nombre") if sede_item else ""),
            "codigo": request.form.get("codigo"),
            "nombre": request.form.get("nombre"),
            "piso": request.form.get("piso"),
            "descripcion": request.form.get("descripcion"),
            "estado": request.form.get("estado"),
        }
        if not sede_id:
            flash("Debes seleccionar una sede.", "warning")
            return render_template("hc/configuracion/consultorios_form.html", modo="crear", item=data, sedes=sedes)
        if not (data["codigo"] or "").strip():
            flash("El código es obligatorio.", "warning")
            return render_template("hc/configuracion/consultorios_form.html", modo="crear", item=data, sedes=sedes)
        if not (data["nombre"] or "").strip():
            flash("El nombre es obligatorio.", "warning")
            return render_template("hc/configuracion/consultorios_form.html", modo="crear", item=data, sedes=sedes)
        if repo_cons.existe_codigo(data["codigo"]):
            flash("Ya existe un consultorio con ese código.", "warning")
            return render_template("hc/configuracion/consultorios_form.html", modo="crear", item=data, sedes=sedes)
        repo_cons.crear(data)
        flash("Consultorio creado correctamente.", "success")
        return redirect(url_for("hc_configuracion.consultorios"))
    return render_template("hc/configuracion/consultorios_form.html", modo="crear", item={}, sedes=sedes)


@bp_hc_configuracion.route("/consultorios/editar/<int:item_id>", methods=["GET", "POST"])
def consultorio_editar(item_id):
    item = repo_cons.obtener(item_id)
    if not item:
        flash("El consultorio no existe.", "error")
        return redirect(url_for("hc_configuracion.consultorios"))
    sedes = repo_sedes.listar()
    if request.method == "POST":
        sede_id_raw = (request.form.get("sede_id") or "").strip()
        sede_id = int(sede_id_raw) if sede_id_raw.isdigit() else None
        sede_item = repo_sedes.obtener(sede_id) if sede_id else None
        data = {
            "sede_id": sede_id,
            "sede_nombre": (sede_item.get("nombre") if sede_item else ""),
            "codigo": request.form.get("codigo"),
            "nombre": request.form.get("nombre"),
            "piso": request.form.get("piso"),
            "descripcion": request.form.get("descripcion"),
            "estado": request.form.get("estado"),
        }
        if not sede_id:
            flash("Debes seleccionar una sede.", "warning")
            return render_template("hc/configuracion/consultorios_form.html", modo="editar", item={**item, **data}, sedes=sedes)
        if not (data["codigo"] or "").strip():
            flash("El código es obligatorio.", "warning")
            return render_template("hc/configuracion/consultorios_form.html", modo="editar", item={**item, **data}, sedes=sedes)
        if not (data["nombre"] or "").strip():
            flash("El nombre es obligatorio.", "warning")
            return render_template("hc/configuracion/consultorios_form.html", modo="editar", item={**item, **data}, sedes=sedes)
        if repo_cons.existe_codigo(data["codigo"], exclude_id=item_id):
            flash("Ya existe otro consultorio con ese código.", "warning")
            return render_template("hc/configuracion/consultorios_form.html", modo="editar", item={**item, **data}, sedes=sedes)
        repo_cons.actualizar(item_id, data)
        flash("Consultorio actualizado correctamente.", "success")
        return redirect(url_for("hc_configuracion.consultorios"))
    return render_template("hc/configuracion/consultorios_form.html", modo="editar", item=item, sedes=sedes)


@bp_hc_configuracion.route("/consultorios/toggle/<int:item_id>", methods=["POST"])
def consultorio_toggle(item_id):
    item = repo_cons.obtener(item_id)
    if not item:
        flash("El consultorio no existe.", "error")
        return redirect(url_for("hc_configuracion.consultorios"))
    nuevo_estado = "INACTIVO" if item.get("estado") == "ACTIVO" else "ACTIVO"
    repo_cons.cambiar_estado(item_id, nuevo_estado)
    flash(f"Consultorio {'inactivado' if nuevo_estado == 'INACTIVO' else 'activado'} correctamente.", "success")
    return redirect(url_for("hc_configuracion.consultorios"))


# ══════════════════════════════════════════════════════════════════
#  PROFESIONALES
# ══════════════════════════════════════════════════════════════════

@bp_hc_configuracion.route("/profesionales")
def profesionales():
    items = repo_prof.listar()
    return render_template("hc/configuracion/profesionales.html", items=items)


@bp_hc_configuracion.route("/profesionales/nuevo", methods=["GET", "POST"])
def profesional_nuevo():
    tipos_documento = repo_td.listar()
    especialidades  = repo_esp.listar()
    sedes           = repo_sedes.listar()
    consultorios    = repo_cons.listar()
    if request.method == "POST":
        data = {
            "tipo_documento_id":    request.form.get("tipo_documento_id"),
            "numero_documento":     request.form.get("numero_documento"),
            "nombres":              request.form.get("nombres"),
            "apellidos":            request.form.get("apellidos"),
            "especialidad_id":      request.form.get("especialidad_id"),
            "sede_id":              request.form.get("sede_id"),
            "consultorio_id":       request.form.get("consultorio_id"),
            "registro_profesional": request.form.get("registro_profesional"),
            "correo":               request.form.get("correo"),
            "telefono":             request.form.get("telefono"),
            "estado":               request.form.get("estado"),
        }
        ctx = dict(modo="crear", item=data, tipos_documento=tipos_documento,
                   especialidades=especialidades, sedes=sedes, consultorios=consultorios)
        if not (data["tipo_documento_id"] or "").strip():
            flash("Debes seleccionar tipo de documento.", "warning")
            return render_template("hc/configuracion/profesionales_form.html", **ctx)
        if not (data["numero_documento"] or "").strip():
            flash("El número de documento es obligatorio.", "warning")
            return render_template("hc/configuracion/profesionales_form.html", **ctx)
        if not (data["nombres"] or "").strip():
            flash("Los nombres son obligatorios.", "warning")
            return render_template("hc/configuracion/profesionales_form.html", **ctx)
        if not (data["apellidos"] or "").strip():
            flash("Los apellidos son obligatorios.", "warning")
            return render_template("hc/configuracion/profesionales_form.html", **ctx)
        if repo_prof.existe_documento(data["tipo_documento_id"], data["numero_documento"]):
            flash("Ya existe un profesional con ese tipo y número de documento.", "warning")
            return render_template("hc/configuracion/profesionales_form.html", **ctx)
        repo_prof.crear(data)
        flash("Profesional creado correctamente.", "success")
        return redirect(url_for("hc_configuracion.profesionales"))
    return render_template("hc/configuracion/profesionales_form.html",
                           modo="crear", item={}, tipos_documento=tipos_documento,
                           especialidades=especialidades, sedes=sedes, consultorios=consultorios)


@bp_hc_configuracion.route("/profesionales/editar/<int:item_id>", methods=["GET", "POST"])
def profesional_editar(item_id):
    item = repo_prof.obtener(item_id)
    if not item:
        flash("El profesional no existe.", "error")
        return redirect(url_for("hc_configuracion.profesionales"))
    tipos_documento = repo_td.listar()
    especialidades  = repo_esp.listar()
    sedes           = repo_sedes.listar()
    consultorios    = repo_cons.listar()
    if request.method == "POST":
        data = {
            "tipo_documento_id":    request.form.get("tipo_documento_id"),
            "numero_documento":     request.form.get("numero_documento"),
            "nombres":              request.form.get("nombres"),
            "apellidos":            request.form.get("apellidos"),
            "especialidad_id":      request.form.get("especialidad_id"),
            "sede_id":              request.form.get("sede_id"),
            "consultorio_id":       request.form.get("consultorio_id"),
            "registro_profesional": request.form.get("registro_profesional"),
            "correo":               request.form.get("correo"),
            "telefono":             request.form.get("telefono"),
            "estado":               request.form.get("estado"),
        }
        if repo_prof.existe_documento(data["tipo_documento_id"], data["numero_documento"], exclude_id=item_id):
            flash("Ya existe otro profesional con ese tipo y número de documento.", "warning")
            return render_template("hc/configuracion/profesionales_form.html",
                                   modo="editar", item={**item, **data},
                                   tipos_documento=tipos_documento, especialidades=especialidades,
                                   sedes=sedes, consultorios=consultorios)
        repo_prof.actualizar(item_id, data)
        flash("Profesional actualizado correctamente.", "success")
        return redirect(url_for("hc_configuracion.profesionales"))
    return render_template("hc/configuracion/profesionales_form.html",
                           modo="editar", item=item, tipos_documento=tipos_documento,
                           especialidades=especialidades, sedes=sedes, consultorios=consultorios)


@bp_hc_configuracion.route("/profesionales/toggle/<int:item_id>", methods=["POST"])
def profesional_toggle(item_id):
    item = repo_prof.obtener(item_id)
    if not item:
        flash("El profesional no existe.", "error")
        return redirect(url_for("hc_configuracion.profesionales"))
    nuevo_estado = "INACTIVO" if item.get("estado") == "ACTIVO" else "ACTIVO"
    repo_prof.cambiar_estado(item_id, nuevo_estado)
    flash(f"Profesional {'inactivado' if nuevo_estado == 'INACTIVO' else 'activado'} correctamente.", "success")
    return redirect(url_for("hc_configuracion.profesionales"))


@bp_hc_configuracion.route("/profesionales/<int:prof_id>/programacion", methods=["GET"])
def prof_programacion_listar(prof_id):
    data = repo_prog.listar_por_profesional(prof_id)
    return jsonify(data)

@bp_hc_configuracion.route("/profesionales/<int:prof_id>/programacion/agregar", methods=["POST"])
def prof_programacion_agregar(prof_id):
    body = request.get_json()
    try:
        resultado = repo_prog.agregar_bloque(
            profesional_id=prof_id,
            dia_semana=int(body["dia_semana"]),
            hora_inicio=body["hora_inicio"],
            hora_fin=body["hora_fin"],
        )
        return jsonify({"ok": True, "data": resultado})
    except ValueError as e:
        return jsonify({"ok": False, "msg": str(e)}), 400
    except Exception as e:
        return jsonify({"ok": False, "msg": str(e)}), 400
    

    
@bp_hc_configuracion.route("/profesionales/programacion/<int:bloque_id>/eliminar", methods=["POST"])
def prof_programacion_eliminar(bloque_id):
    repo_prog.eliminar_bloque(bloque_id)
    return jsonify({"ok": True})


@bp_hc_configuracion.route("/profesionales/<int:prof_id>/bloqueos", methods=["GET"])
def prof_bloqueos_listar(prof_id):
    data = repo_prog.listar_bloqueos(prof_id)
    return jsonify(data)
 
 
@bp_hc_configuracion.route("/profesionales/<int:prof_id>/bloqueos/agregar", methods=["POST"])
def prof_bloqueos_agregar(prof_id):
    body = request.get_json()
    try:
        resultado = repo_prog.agregar_bloqueo(
            profesional_id=prof_id,
            fecha_inicio=body["fecha_inicio"],
            fecha_fin=body["fecha_fin"],
            motivo=body.get("motivo"),
            hora_inicio=body.get("hora_inicio"),  # NUEVO — puede ser null
            hora_fin=body.get("hora_fin"),         # NUEVO — puede ser null
        )
        return jsonify({"ok": True, "data": resultado})
    except Exception as e:
        return jsonify({"ok": False, "msg": str(e)}), 400
 
 
@bp_hc_configuracion.route("/profesionales/bloqueos/<int:bloqueo_id>/eliminar", methods=["POST"])
def prof_bloqueos_eliminar(bloqueo_id):
    repo_prog.eliminar_bloqueo(bloqueo_id)
    return jsonify({"ok": True})



@bp_hc_configuracion.route("/profesionales/<int:prof_id>/disponibilidad", methods=["GET"])
def prof_disponibilidad(prof_id):
    """
    GET /hc/configuracion/profesionales/<id>/disponibilidad?fecha=2026-05-15
    Retorna los rangos horarios donde el profesional puede atender.
    """
    fecha = request.args.get("fecha")
    if not fecha:
        return jsonify({"ok": False, "msg": "Falta el parámetro fecha"}), 400
    try:
        rangos = repo_prog.obtener_disponibilidad(prof_id, fecha)
        return jsonify({"ok": True, "data": rangos})
    except Exception as e:
        return jsonify({"ok": False, "msg": str(e)}), 400

@bp_hc_configuracion.route("/profesionales/<int:prof_id>/alertas", methods=["GET"])
def prof_alertas(prof_id):
    """
    GET /hc/configuracion/profesionales/<id>/alertas?fecha=2026-05-15
    Retorna bloqueos parciales y totales para mostrar en la agenda.
    """
    fecha = request.args.get("fecha")
    if not fecha:
        return jsonify({"ok": False, "msg": "Falta el parámetro fecha"}), 400
    try:
        alertas = repo_prog.obtener_alertas_fecha(prof_id, fecha)
        return jsonify({"ok": True, "data": alertas})
    except Exception as e:
        return jsonify({"ok": False, "msg": str(e)}), 400

@bp_hc_configuracion.route("/profesionales/<int:prof_id>/siguiente-disponible", methods=["GET"])
def prof_siguiente_disponible(prof_id):
    fecha = request.args.get("fecha")
    duracion = int(request.args.get("duracion", 20))
    if not fecha:
        return jsonify({"ok": False, "msg": "Falta fecha"}), 400
    try:
        resultado = repo_prog.buscar_siguiente_disponible(prof_id, fecha, duracion)
        return jsonify({"ok": True, "data": resultado})
    except Exception as e:
        return jsonify({"ok": False, "msg": str(e)}), 400

# ══════════════════════════════════════════════════════════════════
#  SERVICIOS
# ══════════════════════════════════════════════════════════════════

@bp_hc_configuracion.route("/servicios")
def servicios():
    servicios = hc_servicios_repo.listar_servicios()
    return render_template("hc/configuracion/servicios.html", servicios=servicios)


@bp_hc_configuracion.route("/servicios/nuevo")
def servicios_nuevo():
    especialidades = repo_esp.listar()
    return render_template("hc/configuracion/servicios_form.html",
                           modo="crear", servicio=None, especialidades=especialidades)


@bp_hc_configuracion.route("/servicios/crear", methods=["POST"])
def servicios_crear():
    data = {
        "codigo":          request.form.get("codigo"),
        "nombre":          request.form.get("nombre"),
        "especialidad_id": request.form.get("especialidad_id") or None,
        "descripcion":     request.form.get("descripcion"),
    }
    hc_servicios_repo.crear_servicio(data)
    flash("Servicio creado correctamente", "success")
    return redirect("/hc/configuracion/servicios")


@bp_hc_configuracion.route("/servicios/<int:id>/editar")
def servicios_editar(id):
    servicio       = hc_servicios_repo.obtener_servicio(id)
    especialidades = repo_esp.listar()
    return render_template("hc/configuracion/servicios_form.html",
                           modo="editar", servicio=servicio, especialidades=especialidades)


@bp_hc_configuracion.route("/servicios/<int:id>/actualizar", methods=["POST"])
def servicios_actualizar(id):
    data = {
        "codigo":          request.form.get("codigo"),
        "nombre":          request.form.get("nombre"),
        "especialidad_id": request.form.get("especialidad_id") or None,
        "descripcion":     request.form.get("descripcion"),
    }
    hc_servicios_repo.actualizar_servicio(id, data)
    flash("Servicio actualizado", "success")
    return redirect("/hc/configuracion/servicios")


@bp_hc_configuracion.route("/servicios/toggle/<int:item_id>", methods=["POST"])
def servicio_toggle(item_id):
    item = hc_servicios_repo.obtener_servicio(item_id)
    if not item:
        flash("El servicio no existe.", "error")
        return redirect(url_for("hc_configuracion.servicios"))
    nuevo_estado = "INACTIVO" if item.get("estado") == "ACTIVO" else "ACTIVO"
    hc_servicios_repo.cambiar_estado(item_id, nuevo_estado)
    flash(f"Servicio {'inactivado' if nuevo_estado == 'INACTIVO' else 'activado'} correctamente.", "success")
    return redirect(url_for("hc_configuracion.servicios"))


# ══════════════════════════════════════════════════════════════════
#  PAÍSES
# ══════════════════════════════════════════════════════════════════

@bp_hc_configuracion.route("/paises")
def paises():
    items = repo_paises.listar()
    return render_template("hc/configuracion/paises.html", items=items)


@bp_hc_configuracion.route("/paises/nuevo", methods=["GET", "POST"])
def pais_nuevo():
    if request.method == "POST":
        data = {
            "codigo_iso2": (request.form.get("codigo_iso2") or "").strip().upper(),
            "codigo_iso3": (request.form.get("codigo_iso3") or "").strip().upper(),
            "nombre":      (request.form.get("nombre") or "").strip(),
        }
        if not data["codigo_iso2"]:
            flash("El código ISO2 es obligatorio.", "warning")
            return render_template("hc/configuracion/paises_form.html", modo="crear", item=data)
        if not data["nombre"]:
            flash("El nombre del país es obligatorio.", "warning")
            return render_template("hc/configuracion/paises_form.html", modo="crear", item=data)
        existente = repo_paises.obtener_por_iso2(data["codigo_iso2"])
        if existente:
            flash(f"Ya existe un país con código ISO2 '{data['codigo_iso2']}'.", "warning")
            return render_template("hc/configuracion/paises_form.html", modo="crear", item=data)
        try:
            repo_paises.crear(data)
            flash("País creado correctamente.", "success")
            return redirect(url_for("hc_configuracion.paises"))
        except Exception:
            flash("Error al crear el país.", "error")
            return render_template("hc/configuracion/paises_form.html", modo="crear", item=data)
    return render_template("hc/configuracion/paises_form.html", modo="crear", item={})


@bp_hc_configuracion.route("/paises/editar/<int:item_id>", methods=["GET", "POST"])
def pais_editar(item_id):
    item = repo_paises.obtener(item_id)
    if not item:
        flash("El país no existe.", "error")
        return redirect(url_for("hc_configuracion.paises"))
    if request.method == "POST":
        data = {
            "codigo_iso2": (request.form.get("codigo_iso2") or "").strip().upper(),
            "codigo_iso3": (request.form.get("codigo_iso3") or "").strip().upper(),
            "nombre":      (request.form.get("nombre") or "").strip(),
        }
        if not data["codigo_iso2"]:
            flash("El código ISO2 es obligatorio.", "warning")
            return render_template("hc/configuracion/paises_form.html", modo="editar", item={**item, **data})
        if not data["nombre"]:
            flash("El nombre del país es obligatorio.", "warning")
            return render_template("hc/configuracion/paises_form.html", modo="editar", item={**item, **data})
        existente = repo_paises.obtener_por_iso2(data["codigo_iso2"])
        if existente and existente["id"] != item_id:
            flash(f"Ya existe otro país con ISO2 '{data['codigo_iso2']}'.", "warning")
            return render_template("hc/configuracion/paises_form.html", modo="editar", item={**item, **data})
        try:
            repo_paises.actualizar(item_id, data)
            flash("País actualizado correctamente.", "success")
            return redirect(url_for("hc_configuracion.paises"))
        except Exception:
            flash("Error al actualizar el país.", "error")
            return render_template("hc/configuracion/paises_form.html", modo="editar", item={**item, **data})
    return render_template("hc/configuracion/paises_form.html", modo="editar", item=item)


@bp_hc_configuracion.route("/paises/toggle/<int:item_id>", methods=["POST"])
def pais_toggle(item_id):
    item = repo_paises.obtener(item_id)
    if not item:
        flash("El país no existe.", "error")
        return redirect(url_for("hc_configuracion.paises"))
    nuevo_estado = "INACTIVO" if item.get("estado") == "ACTIVO" else "ACTIVO"
    repo_paises.cambiar_estado(item_id, nuevo_estado)
    flash(f"País {'inactivado' if nuevo_estado == 'INACTIVO' else 'activado'} correctamente.", "success")
    return redirect(url_for("hc_configuracion.paises"))


# ══════════════════════════════════════════════════════════════════
#  CIE10
# ══════════════════════════════════════════════════════════════════

@bp_hc_configuracion.route("/cie10")
def cie10():
    items = repo_cie10.listar()
    return render_template("hc/configuracion/cie10.html", items=items)


@bp_hc_configuracion.route("/cie10/nuevo")
def cie10_nuevo():
    return render_template("hc/configuracion/cie10_form.html", modo="crear", item=None)


@bp_hc_configuracion.route("/cie10/crear", methods=["POST"])
def cie10_crear():
    data = {
        "codigo":      request.form.get("codigo"),
        "nombre":      request.form.get("nombre"),
        "descripcion": request.form.get("descripcion"),
        "categoria":   request.form.get("categoria"),
    }
    repo_cie10.crear(data)
    flash("Diagnóstico CIE10 creado", "success")
    return redirect("/hc/configuracion/cie10")


@bp_hc_configuracion.route("/cie10/<int:item_id>/editar")
def cie10_editar(item_id):
    item = repo_cie10.obtener(item_id)
    return render_template("hc/configuracion/cie10_form.html", modo="editar", item=item)


@bp_hc_configuracion.route("/cie10/<int:item_id>/actualizar", methods=["POST"])
def cie10_actualizar(item_id):
    data = {
        "codigo":      request.form.get("codigo"),
        "nombre":      request.form.get("nombre"),
        "descripcion": request.form.get("descripcion"),
        "categoria":   request.form.get("categoria"),
    }
    repo_cie10.actualizar(item_id, data)
    flash("Diagnóstico actualizado", "success")
    return redirect("/hc/configuracion/cie10")


@bp_hc_configuracion.route("/cie10/toggle/<int:item_id>", methods=["POST"])
def cie10_toggle(item_id):
    item = repo_cie10.obtener(item_id)
    if not item:
        flash("El diagnóstico no existe.", "error")
        return redirect(url_for("hc_configuracion.cie10"))
    nuevo_estado = "INACTIVO" if item.get("estado") == "ACTIVO" else "ACTIVO"
    repo_cie10.cambiar_estado(item_id, nuevo_estado)
    flash(f"Diagnóstico {'inactivado' if nuevo_estado == 'INACTIVO' else 'activado'} correctamente.", "success")
    return redirect(url_for("hc_configuracion.cie10"))


@bp_hc_configuracion.route("/cie10/buscar")
def cie10_buscar():
    q = request.args.get("q", "")
    items = repo_cie10.buscar(q)
    return jsonify(items)


# ══════════════════════════════════════════════════════════════════
#  DEPARTAMENTOS
# ══════════════════════════════════════════════════════════════════

@bp_hc_configuracion.route("/departamentos")
def departamentos():
    items = repo_dep.listar()
    return render_template("hc/configuracion/departamentos.html", items=items)


@bp_hc_configuracion.route("/departamentos/nuevo", methods=["GET", "POST"])
def departamento_nuevo():
    paises = repo_paises.listar()
    if request.method == "POST":
        data = {
            "pais_id": request.form.get("pais_id"),
            "codigo":  request.form.get("codigo"),
            "nombre":  request.form.get("nombre"),
        }
        repo_dep.crear(data)
        flash("Departamento creado correctamente.", "success")
        return redirect(url_for("hc_configuracion.departamentos"))
    return render_template("hc/configuracion/departamentos_form.html",
                           paises=paises, item={}, modo="crear")


@bp_hc_configuracion.route("/departamentos/editar/<int:item_id>", methods=["GET", "POST"])
def departamento_editar(item_id):
    item   = repo_dep.obtener(item_id)
    paises = repo_paises.listar()
    if request.method == "POST":
        data = {
            "pais_id": request.form.get("pais_id"),
            "codigo":  request.form.get("codigo"),
            "nombre":  request.form.get("nombre"),
        }
        repo_dep.actualizar(item_id, data)
        flash("Departamento actualizado.", "success")
        return redirect(url_for("hc_configuracion.departamentos"))
    return render_template("hc/configuracion/departamentos_form.html",
                           item=item, paises=paises, modo="editar")


@bp_hc_configuracion.route("/departamentos/toggle/<int:item_id>", methods=["POST"])
def departamento_toggle(item_id):
    item = repo_dep.obtener(item_id)
    nuevo_estado = "INACTIVO" if item["estado"] == "ACTIVO" else "ACTIVO"
    repo_dep.cambiar_estado(item_id, nuevo_estado)
    flash("Estado actualizado.", "success")
    return redirect(url_for("hc_configuracion.departamentos"))


# ══════════════════════════════════════════════════════════════════
#  MUNICIPIOS
# ══════════════════════════════════════════════════════════════════

@bp_hc_configuracion.route("/municipios")
def municipios():
    items = repo_muni.listar()
    return render_template("hc/configuracion/municipios.html", items=items)


@bp_hc_configuracion.route("/municipios/nuevo", methods=["GET", "POST"])
def municipio_nuevo():
    departamentos = repo_dep.listar()
    if request.method == "POST":
        data = {
            "departamento_id": request.form.get("departamento_id"),
            "codigo":          request.form.get("codigo"),
            "nombre":          request.form.get("nombre"),
        }
        repo_muni.crear(data)
        flash("Municipio creado correctamente.", "success")
        return redirect(url_for("hc_configuracion.municipios"))
    return render_template("hc/configuracion/municipios_form.html", departamentos=departamentos)


@bp_hc_configuracion.route("/municipios/editar/<int:item_id>", methods=["GET", "POST"])
def municipio_editar(item_id):
    item          = repo_muni.obtener(item_id)
    departamentos = repo_dep.listar()
    if request.method == "POST":
        data = {
            "departamento_id": request.form.get("departamento_id"),
            "codigo":          request.form.get("codigo"),
            "nombre":          request.form.get("nombre"),
        }
        repo_muni.actualizar(item_id, data)
        flash("Municipio actualizado.", "success")
        return redirect(url_for("hc_configuracion.municipios"))
    return render_template("hc/configuracion/municipios_form.html",
                           item=item, departamentos=departamentos, modo="editar")


@bp_hc_configuracion.route("/municipios/toggle/<int:item_id>", methods=["POST"])
def municipio_toggle(item_id):
    item = repo_muni.obtener(item_id)
    nuevo_estado = "INACTIVO" if item["estado"] == "ACTIVO" else "ACTIVO"
    repo_muni.cambiar_estado(item_id, nuevo_estado)
    flash("Estado actualizado.", "success")
    return redirect(url_for("hc_configuracion.municipios"))


# ══════════════════════════════════════════════════════════════════
#  EPS
# ══════════════════════════════════════════════════════════════════

@bp_hc_configuracion.route("/eps")
def eps():
    items = repo_eps.listar()
    return render_template("hc/configuracion/eps.html", items=items)


@bp_hc_configuracion.route("/eps/nuevo", methods=["GET", "POST"])
def eps_nuevo():
    if request.method == "POST":
        data = {
            "codigo":  request.form.get("codigo"),
            "nombre":  request.form.get("nombre"),
            "nit":     request.form.get("nit"),
            "regimen": request.form.get("regimen"),
        }
        repo_eps.crear(data)
        flash("EPS creada correctamente.", "success")
        return redirect(url_for("hc_configuracion.eps"))
    return render_template("hc/configuracion/eps_form.html", item={}, modo="crear")


@bp_hc_configuracion.route("/eps/editar/<int:item_id>", methods=["GET", "POST"])
def eps_editar(item_id):
    item = repo_eps.obtener(item_id)
    if request.method == "POST":
        data = {
            "codigo":  request.form.get("codigo"),
            "nombre":  request.form.get("nombre"),
            "nit":     request.form.get("nit"),
            "regimen": request.form.get("regimen"),
        }
        repo_eps.actualizar(item_id, data)
        flash("EPS actualizada.", "success")
        return redirect(url_for("hc_configuracion.eps"))
    return render_template("hc/configuracion/eps_form.html", item=item, modo="editar")


@bp_hc_configuracion.route("/eps/toggle/<int:item_id>", methods=["POST"])
def eps_toggle(item_id):
    item = repo_eps.obtener(item_id)
    nuevo_estado = "INACTIVO" if item["estado"] == "ACTIVO" else "ACTIVO"
    repo_eps.cambiar_estado(item_id, nuevo_estado)
    flash("Estado actualizado.", "success")
    return redirect(url_for("hc_configuracion.eps"))


@bp_hc_configuracion.route("/eps/buscar")
def eps_buscar():
    q = request.args.get("q", "")
    items = repo_eps.buscar(q)
    return jsonify(items)


# ══════════════════════════════════════════════════════════════════
#  RECURSOS
# ══════════════════════════════════════════════════════════════════

@bp_hc_configuracion.route("/recursos")
def recursos_listar():
    data = repo_recursos.listar()
    return render_template("hc/configuracion/recursos.html", data=data)


@bp_hc_configuracion.route("/recursos/nuevo", methods=["GET", "POST"])
def recursos_nuevo():
    if request.method == "POST":
        data = {
            "codigo":         request.form.get("codigo"),
            "nombre":         request.form.get("nombre"),
            "tipo":           request.form.get("tipo"),
            "descripcion":    request.form.get("descripcion"),
            "sede_id":        request.form.get("sede_id") or None,
            "consultorio_id": request.form.get("consultorio_id") or None,
        }
        repo_recursos.crear(data)
        return redirect(url_for("hc_configuracion.recursos_listar"))
    recurso = {"codigo": "", "nombre": "", "tipo": "", "descripcion": "", "sede_id": None, "consultorio_id": None}
    return render_template("hc/configuracion/recursos_form.html",
                           recurso=recurso, modo="crear",
                           sedes=repo_sedes.listar(), consultorios=repo_cons.listar())


@bp_hc_configuracion.route("/recursos/editar/<int:id>", methods=["GET", "POST"])
def recursos_editar(id):
    recurso = repo_recursos.obtener(id)
    if request.method == "POST":
        data = {
            "codigo":         request.form.get("codigo"),
            "nombre":         request.form.get("nombre"),
            "tipo":           request.form.get("tipo"),
            "descripcion":    request.form.get("descripcion"),
            "sede_id":        request.form.get("sede_id") or None,
            "consultorio_id": request.form.get("consultorio_id") or None,
        }
        repo_recursos.actualizar(id, data)
        return redirect(url_for("hc_configuracion.recursos_listar"))
    return render_template("hc/configuracion/recursos_form.html",
                           recurso=recurso, modo="editar",
                           sedes=repo_sedes.listar(), consultorios=repo_cons.listar())


@bp_hc_configuracion.route("/recursos/toggle/<int:id>", methods=["POST"])
def recursos_toggle(id):
    repo_recursos.toggle(id)
    return redirect(url_for("hc_configuracion.recursos_listar"))


# ══════════════════════════════════════════════════════════════════
#  AGENDAS
# ══════════════════════════════════════════════════════════════════

@bp_hc_configuracion.route("/agendas")
def agendas():
    data = agendas_repo.listar()
    return render_template("hc/configuracion/agendas.html", data=data)


@bp_hc_configuracion.route("/agendas/nuevo", methods=["GET", "POST"])
def agendas_nuevo():
    if request.method == "POST":
        data = {
            "tipo":           request.form.get("tipo"),
            "profesional_id": request.form.get("profesional_id") or None,
            "recurso_id":     request.form.get("recurso_id") or None,
            "dia_semana":     int(request.form.get("dia_semana")),
            "hora_inicio":    request.form.get("hora_inicio"),
            "hora_fin":       request.form.get("hora_fin"),
            "duracion_min":   int(request.form.get("duracion_min")),
        }
        agendas_repo.crear(data)
        return redirect(url_for("hc_configuracion.agendas"))
    return render_template("hc/configuracion/agendas_form.html", modo="crear")


@bp_hc_configuracion.route("/agendas/toggle/<int:id>")
def agendas_toggle(id):
    agendas_repo.toggle(id)
    return redirect(url_for("hc_configuracion.agendas"))


# ══════════════════════════════════════════════════════════════════
#  CUPS
# ══════════════════════════════════════════════════════════════════

@bp_hc_configuracion.route("/cups")
def cups():
    items = repo_cups.listar()
    return render_template("hc/configuracion/cups.html", items=items)


@bp_hc_configuracion.route("/cups/nuevo")
def cups_nuevo():
    return render_template("hc/configuracion/cups_form.html", modo="crear", item=None)


@bp_hc_configuracion.route("/cups/crear", methods=["POST"])
def cups_crear():
    data = {
        "codigo":      request.form.get("codigo", "").strip().upper(),
        "descripcion": request.form.get("descripcion", "").strip(),
        "estado":      request.form.get("estado", "ACTIVO"),
    }
    repo_cups.crear(data)
    flash("Procedimiento CUPS creado correctamente.", "success")
    return redirect(url_for("hc_configuracion.cups"))


@bp_hc_configuracion.route("/cups/<int:item_id>/editar")
def cups_editar(item_id):
    item = repo_cups.obtener(item_id)
    if not item:
        flash("Procedimiento no encontrado.", "error")
        return redirect(url_for("hc_configuracion.cups"))
    return render_template("hc/configuracion/cups_form.html", modo="editar", item=item)


@bp_hc_configuracion.route("/cups/<int:item_id>/actualizar", methods=["POST"])
def cups_actualizar(item_id):
    data = {
        "codigo":      request.form.get("codigo", "").strip().upper(),
        "descripcion": request.form.get("descripcion", "").strip(),
        "estado":      request.form.get("estado", "ACTIVO"),
    }
    repo_cups.actualizar(item_id, data)
    flash("Procedimiento CUPS actualizado.", "success")
    return redirect(url_for("hc_configuracion.cups"))


@bp_hc_configuracion.route("/cups/toggle/<int:item_id>", methods=["POST"])
def cups_toggle(item_id):
    item = repo_cups.obtener(item_id)
    if not item:
        flash("El procedimiento no existe.", "error")
        return redirect(url_for("hc_configuracion.cups"))
    nuevo_estado = "INACTIVO" if item.get("estado") == "ACTIVO" else "ACTIVO"
    repo_cups.cambiar_estado(item_id, nuevo_estado)
    flash(f"Procedimiento {'inactivado' if nuevo_estado == 'INACTIVO' else 'activado'} correctamente.", "success")
    return redirect(url_for("hc_configuracion.cups"))


@bp_hc_configuracion.route("/cups/buscar")
def cups_buscar():
    q = request.args.get("q", "").strip()
    items = repo_cups.buscar(q) if q else []
    return jsonify(items)


@bp_hc_configuracion.route("/cups/exportar/csv")
def cups_exportar_csv():
    items = repo_cups.listar_todos_exportar()
    def generate():
        yield ",".join(["ID", "Código", "Descripción", "Estado"]) + "\n"
        for row in items:
            yield ",".join([
                str(row.get("id", "")),
                row.get("codigo", ""),
                '"' + row.get("descripcion", "").replace('"', '""') + '"',
                row.get("estado", ""),
            ]) + "\n"
    return Response(generate(), mimetype="text/csv",
                    headers={"Content-Disposition": "attachment; filename=cups_export.csv"})


@bp_hc_configuracion.route("/cups/exportar/json")
def cups_exportar_json():
    items = repo_cups.listar_todos_exportar()
    return jsonify(items)


@bp_hc_configuracion.route("/cups/importar", methods=["POST"])
def cups_importar():
    archivo = request.files.get("archivo")
    if not archivo:
        return jsonify({"ok": False, "msg": "No se recibió archivo"}), 400
    nombre    = archivo.filename.lower()
    registros = []
    if nombre.endswith(".csv"):
        stream = io.StringIO(archivo.stream.read().decode("utf-8-sig"))
        reader = csv.DictReader(stream)
        for fila in reader:
            codigo      = (fila.get("codigo") or fila.get("Código") or "").strip().upper()
            descripcion = (fila.get("descripcion") or fila.get("Descripción") or "").strip()
            estado      = (fila.get("estado") or fila.get("Estado") or "ACTIVO").strip().upper()
            if not codigo or not descripcion:
                continue
            if estado not in ("ACTIVO", "INACTIVO"):
                estado = "ACTIVO"
            registros.append({"codigo": codigo, "descripcion": descripcion, "estado": estado})
    elif nombre.endswith(".xlsx") or nombre.endswith(".xls"):
        try:
            import openpyxl
            wb   = openpyxl.load_workbook(archivo.stream, read_only=True, data_only=True)
            ws   = wb.active
            rows = list(ws.iter_rows(values_only=True))
        except Exception as exc:
            return jsonify({"ok": False, "msg": f"No se pudo leer el Excel: {exc}"}), 400
        if not rows:
            return jsonify({"ok": False, "msg": "El archivo está vacío"}), 400
        headers = [str(h).strip().lower() if h else "" for h in rows[0]]
        for fila in rows[1:]:
            row_dict    = dict(zip(headers, fila))
            codigo      = str(row_dict.get("codigo") or "").strip().upper()
            descripcion = str(row_dict.get("descripcion") or "").strip()
            estado      = str(row_dict.get("estado") or "ACTIVO").strip().upper()
            if not codigo or not descripcion:
                continue
            if estado not in ("ACTIVO", "INACTIVO"):
                estado = "ACTIVO"
            registros.append({"codigo": codigo, "descripcion": descripcion, "estado": estado})
    else:
        return jsonify({"ok": False, "msg": "Formato no soportado. Use .csv o .xlsx"}), 400
    if not registros:
        return jsonify({"ok": False, "msg": "No se encontraron registros válidos en el archivo"}), 400
    repo_cups.importar_lote(registros)
    return jsonify({"ok": True, "importados": len(registros)})


# ══════════════════════════════════════════════════════════════════
#  PROCEDIMIENTOS DE PROFESIONAL (AJAX)
# ══════════════════════════════════════════════════════════════════

@bp_hc_configuracion.route("/profesionales/<int:prof_id>/procedimientos", methods=["GET"])
def prof_procedimientos_listar(prof_id):
    return jsonify(prof_proc_repo.listar_por_profesional(prof_id))


@bp_hc_configuracion.route("/profesionales/<int:prof_id>/procedimientos/agregar", methods=["POST"])
def prof_procedimientos_agregar(prof_id):
    body = request.get_json()
    try:
        prof_proc_repo.agregar(
            profesional_id=prof_id,
            cups_id=int(body["cups_id"]),
            duracion_min=int(body.get("duracion_min", 20)),
        )
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "msg": str(e)}), 400


@bp_hc_configuracion.route("/profesionales/procedimientos/<int:id>/duracion", methods=["POST"])
def prof_procedimientos_duracion(id):
    body = request.get_json()
    prof_proc_repo.actualizar_duracion(id, int(body["duracion_min"]))
    return jsonify({"ok": True})


@bp_hc_configuracion.route("/profesionales/procedimientos/<int:id>/eliminar", methods=["POST"])
def prof_procedimientos_eliminar(id):
    prof_proc_repo.eliminar(id)
    return jsonify({"ok": True})


# ══════════════════════════════════════════════════════════════════
#  PROCEDIMIENTOS DE RECURSO (AJAX)
# ══════════════════════════════════════════════════════════════════

@bp_hc_configuracion.route("/recursos/<int:rec_id>/procedimientos", methods=["GET"])
def rec_procedimientos_listar(rec_id):
    return jsonify(rec_proc_repo.listar_por_recurso(rec_id))


@bp_hc_configuracion.route("/recursos/<int:rec_id>/procedimientos/agregar", methods=["POST"])
def rec_procedimientos_agregar(rec_id):
    body = request.get_json()
    try:
        rec_proc_repo.agregar(
            recurso_id=rec_id,
            cups_id=int(body["cups_id"]),
            duracion_min=int(body.get("duracion_min", 20)),
        )
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "msg": str(e)}), 400


@bp_hc_configuracion.route("/recursos/procedimientos/<int:id>/duracion", methods=["POST"])
def rec_procedimientos_duracion(id):
    body = request.get_json()
    rec_proc_repo.actualizar_duracion(id, int(body["duracion_min"]))
    return jsonify({"ok": True})


@bp_hc_configuracion.route("/recursos/procedimientos/<int:id>/eliminar", methods=["POST"])
def rec_procedimientos_eliminar(id):
    rec_proc_repo.eliminar(id)
    return jsonify({"ok": True})


@bp_hc_configuracion.route("/cups/buscar-ajax")
def cups_buscar_ajax():
    q = request.args.get("q", "").strip()
    if len(q) < 2:
        return jsonify([])
    try:
        r = (
            get_supabase_public()
            .table("hc_cups")
            .select("id,codigo,descripcion")
            .or_(f"codigo.ilike.%{q}%,descripcion.ilike.%{q}%")
            .eq("estado", "ACTIVO")
            .limit(20)
            .execute()
        )
        return jsonify(r.data or [])
    except Exception:
        return jsonify([])


# ══════════════════════════════════════════════════════════════════
#  MÉDICOS REMITENTES
# ══════════════════════════════════════════════════════════════════

@bp_hc_configuracion.route("/medicos-remitentes")
def medicos_remitentes():
    items = repo_mr.listar()
    return render_template("hc/configuracion/medicos_remitentes.html", items=items)


@bp_hc_configuracion.route("/medicos-remitentes/nuevo", methods=["GET", "POST"])
def medico_remitente_nuevo():
    tipos_documento = repo_td.listar()
    if request.method == "POST":
        data = {
            "tipo_documento_id": request.form.get("tipo_documento_id"),
            "numero_documento":  request.form.get("numero_documento"),
            "nombres":           request.form.get("nombres"),
            "apellidos":         request.form.get("apellidos"),
            "especialidad":      request.form.get("especialidad"),
            "institucion":       request.form.get("institucion"),
            "telefono":          request.form.get("telefono"),
            "correo":            request.form.get("correo"),
            "estado":            request.form.get("estado"),
        }
        ctx = dict(modo="crear", item=data, tipos_documento=tipos_documento)
        if not (data["tipo_documento_id"] or "").strip():
            flash("Debes seleccionar tipo de documento.", "warning")
            return render_template("hc/configuracion/medicos_remitentes_form.html", **ctx)
        if not (data["numero_documento"] or "").strip():
            flash("El número de documento es obligatorio.", "warning")
            return render_template("hc/configuracion/medicos_remitentes_form.html", **ctx)
        if not (data["nombres"] or "").strip():
            flash("Los nombres son obligatorios.", "warning")
            return render_template("hc/configuracion/medicos_remitentes_form.html", **ctx)
        if not (data["apellidos"] or "").strip():
            flash("Los apellidos son obligatorios.", "warning")
            return render_template("hc/configuracion/medicos_remitentes_form.html", **ctx)
        if repo_mr.existe_documento(data["tipo_documento_id"], data["numero_documento"]):
            flash("Ya existe un médico remitente con ese documento.", "warning")
            return render_template("hc/configuracion/medicos_remitentes_form.html", **ctx)
        repo_mr.crear(data)
        flash("Médico remitente creado correctamente.", "success")
        return redirect(url_for("hc_configuracion.medicos_remitentes"))
    return render_template("hc/configuracion/medicos_remitentes_form.html",
                           modo="crear", item={}, tipos_documento=tipos_documento)


@bp_hc_configuracion.route("/medicos-remitentes/editar/<int:item_id>", methods=["GET", "POST"])
def medico_remitente_editar(item_id):
    item = repo_mr.obtener(item_id)
    if not item:
        flash("El médico remitente no existe.", "error")
        return redirect(url_for("hc_configuracion.medicos_remitentes"))
    tipos_documento = repo_td.listar()
    if request.method == "POST":
        data = {
            "tipo_documento_id": request.form.get("tipo_documento_id"),
            "numero_documento":  request.form.get("numero_documento"),
            "nombres":           request.form.get("nombres"),
            "apellidos":         request.form.get("apellidos"),
            "especialidad":      request.form.get("especialidad"),
            "institucion":       request.form.get("institucion"),
            "telefono":          request.form.get("telefono"),
            "correo":            request.form.get("correo"),
            "estado":            request.form.get("estado"),
        }
        if repo_mr.existe_documento(data["tipo_documento_id"], data["numero_documento"], exclude_id=item_id):
            flash("Ya existe otro médico remitente con ese documento.", "warning")
            return render_template("hc/configuracion/medicos_remitentes_form.html",
                                   modo="editar", item={**item, **data}, tipos_documento=tipos_documento)
        repo_mr.actualizar(item_id, data)
        flash("Médico remitente actualizado correctamente.", "success")
        return redirect(url_for("hc_configuracion.medicos_remitentes"))
    return render_template("hc/configuracion/medicos_remitentes_form.html",
                           modo="editar", item=item, tipos_documento=tipos_documento)


@bp_hc_configuracion.route("/medicos-remitentes/toggle/<int:item_id>", methods=["POST"])
def medico_remitente_toggle(item_id):
    item = repo_mr.obtener(item_id)
    if not item:
        flash("El médico remitente no existe.", "error")
        return redirect(url_for("hc_configuracion.medicos_remitentes"))
    nuevo_estado = "INACTIVO" if item.get("estado") == "ACTIVO" else "ACTIVO"
    repo_mr.cambiar_estado(item_id, nuevo_estado)
    flash(f"Médico remitente {'inactivado' if nuevo_estado == 'INACTIVO' else 'activado'} correctamente.", "success")
    return redirect(url_for("hc_configuracion.medicos_remitentes"))


@bp_hc_configuracion.route("/medicos-remitentes/buscar")
def medico_remitente_buscar():
    q = request.args.get("q", "")
    items = repo_mr.buscar(q)
    return jsonify(items)


# ══════════════════════════════════════════════════════════════════
#  CONSTANTE MANUALES TARIFARIOS (para contratos)
# ══════════════════════════════════════════════════════════════════



# ══════════════════════════════════════════════════════════════════
#  CLIENTES
# ══════════════════════════════════════════════════════════════════

@bp_hc_configuracion.route("/clientes")
def clientes():
    data = repo_clientes.listar()
    return render_template("hc/configuracion/clientes_lista.html", data=data)


@bp_hc_configuracion.route("/clientes/nuevo", methods=["GET", "POST"])
def cliente_nuevo():
    if request.method == "GET":
        return render_template("hc/configuracion/cliente_form.html", modo="crear", cliente={})
    form = request.form
    data = {
        "codigo":        (form.get("codigo") or "").strip().upper(),
        "estado":         form.get("estado") or "ACTIVO",
        "nit":           (form.get("nit") or "").strip() or None,
        "nit_contab":    (form.get("nit_contab") or "").strip() or None,
        "nit_tercero":   (form.get("nit_tercero") or "").strip() or None,
        "nombre":        (form.get("nombre") or "").strip(),
        "direccion":     (form.get("direccion") or "").strip() or None,
        "telefonos":     (form.get("telefonos") or "").strip() or None,
        "contacto":      (form.get("contacto") or "").strip() or None,
        "email":         (form.get("email") or "").strip().lower() or None,
        "cod_prestador": (form.get("cod_prestador") or "").strip() or None,
        "cod_contable":  (form.get("cod_contable") or "").strip() or None,
    }
    if not data["codigo"]:
        flash("El código es obligatorio.", "warning")
        return render_template("hc/configuracion/cliente_form.html", modo="crear", cliente=form)
    if not data["nombre"]:
        flash("El nombre es obligatorio.", "warning")
        return render_template("hc/configuracion/cliente_form.html", modo="crear", cliente=form)
    if repo_clientes.existe_codigo(data["codigo"]):
        flash("Ya existe un cliente con ese código.", "warning")
        return render_template("hc/configuracion/cliente_form.html", modo="crear", cliente=form)
    try:
        repo_clientes.crear(data)
        flash("Cliente creado correctamente.", "success")
        return redirect(url_for("hc_configuracion.clientes"))
    except Exception as e:
        flash(f"Error al guardar el cliente: {e}", "danger")
        return render_template("hc/configuracion/cliente_form.html", modo="crear", cliente=form)


@bp_hc_configuracion.route("/clientes/ver/<int:cliente_id>")
def cliente_ver(cliente_id):
    cliente = repo_clientes.obtener(cliente_id)
    if not cliente:
        flash("El cliente no existe.", "error")
        return redirect(url_for("hc_configuracion.clientes"))
    return render_template("hc/configuracion/cliente_detalle.html", cliente=cliente)


@bp_hc_configuracion.route("/clientes/conteos")
def clientes_conteos():
    ids_raw = request.args.get("ids", "")
    if not ids_raw:
        return jsonify([])
    try:
        ids = [int(i) for i in ids_raw.split(",") if i.strip().isdigit()]
    except ValueError:
        return jsonify([])
    if not ids:
        return jsonify([])
    try:
        resultado = []
        for cid in ids:
            res = (
                get_supabase_public()
                .table("hc_contratos")
                .select("id", count="exact")
                .eq("cliente_id", cid)
                .execute()
            )
            resultado.append({"cliente_id": cid, "total": res.count or 0})
        return jsonify(resultado)
    except Exception as e:
        print("ERROR conteos:", e)
        return jsonify([])


@bp_hc_configuracion.route("/clientes/editar/<int:cliente_id>", methods=["GET", "POST"])
def cliente_editar(cliente_id):
    cliente = repo_clientes.obtener(cliente_id)
    if not cliente:
        flash("El cliente no existe.", "error")
        return redirect(url_for("hc_configuracion.clientes"))
    if request.method == "GET":
        return render_template("hc/configuracion/cliente_form.html", modo="editar", cliente=cliente)
    form = request.form
    data = {
        "codigo":        (form.get("codigo") or "").strip().upper(),
        "estado":         form.get("estado") or "ACTIVO",
        "nit":           (form.get("nit") or "").strip() or None,
        "nit_contab":    (form.get("nit_contab") or "").strip() or None,
        "nit_tercero":   (form.get("nit_tercero") or "").strip() or None,
        "nombre":        (form.get("nombre") or "").strip(),
        "direccion":     (form.get("direccion") or "").strip() or None,
        "telefonos":     (form.get("telefonos") or "").strip() or None,
        "contacto":      (form.get("contacto") or "").strip() or None,
        "email":         (form.get("email") or "").strip().lower() or None,
        "cod_prestador": (form.get("cod_prestador") or "").strip() or None,
        "cod_contable":  (form.get("cod_contable") or "").strip() or None,
    }
    if not data["codigo"]:
        flash("El código es obligatorio.", "warning")
        return render_template("hc/configuracion/cliente_form.html", modo="editar", cliente={**cliente, **form})
    if not data["nombre"]:
        flash("El nombre es obligatorio.", "warning")
        return render_template("hc/configuracion/cliente_form.html", modo="editar", cliente={**cliente, **form})
    if repo_clientes.existe_codigo(data["codigo"], exclude_id=cliente_id):
        flash("Ya existe otro cliente con ese código.", "warning")
        return render_template("hc/configuracion/cliente_form.html", modo="editar", cliente={**cliente, **form})
    try:
        repo_clientes.actualizar(cliente_id, data)
        flash("Cliente actualizado correctamente.", "success")
        return redirect(url_for("hc_configuracion.cliente_editar", cliente_id=cliente_id))
    except Exception as e:
        flash(f"Error al actualizar el cliente: {e}", "danger")
        return render_template("hc/configuracion/cliente_form.html", modo="editar", cliente={**cliente, **form})


@bp_hc_configuracion.route("/clientes/toggle/<int:cliente_id>", methods=["POST"])
def cliente_toggle(cliente_id):
    cliente = repo_clientes.obtener(cliente_id)
    if not cliente:
        flash("El cliente no existe.", "error")
        return redirect(url_for("hc_configuracion.clientes"))
    nuevo = "INACTIVO" if cliente.get("estado") == "ACTIVO" else "ACTIVO"
    repo_clientes.cambiar_estado(cliente_id, nuevo)
    flash(f"Cliente {'inactivado' if nuevo == 'INACTIVO' else 'activado'} correctamente.", "success")
    return redirect(url_for("hc_configuracion.clientes"))


@bp_hc_configuracion.route("/clientes/<int:cliente_id>/contratos")
def cliente_contratos_json(cliente_id):
    data = repo_contratos.listar_por_cliente(cliente_id)
    return jsonify(data)


# ══════════════════════════════════════════════════════════════════
#  CONTRATOS
# ══════════════════════════════════════════════════════════════════

@bp_hc_configuracion.route("/contratos/nuevo", methods=["GET", "POST"])
def contrato_nuevo():
    cliente_id_raw = request.args.get("cliente_id") or request.form.get("cliente_id")
    if not cliente_id_raw:
        flash("Debes seleccionar un cliente.", "warning")
        return redirect(url_for("hc_configuracion.clientes"))
    cliente_id = int(cliente_id_raw)
    cliente    = repo_clientes.obtener(cliente_id)
    if not cliente:
        flash("Cliente no encontrado.", "warning")
        return redirect(url_for("hc_configuracion.clientes"))
    if request.method == "GET":
        return render_template("hc/configuracion/contrato_form.html",
                               modo="crear", contrato={}, cliente=cliente,
                               sedes=repo_sedes.listar(),
                               manuales_tarifarios=repo_manuales.listar_activos())
    form = request.form
    data = _contrato_payload(form, cliente_id)
    if not data["nro_contrato"]:
        flash("El número de contrato es obligatorio.", "warning")
        return render_template("hc/configuracion/contrato_form.html",
                               modo="crear", contrato=form, cliente=cliente,
                               sedes=repo_sedes.listar(),
                               manuales_tarifarios=repo_manuales.listar_activos())
    if repo_contratos.existe_nro(data["nro_contrato"]):
        flash("Ya existe un contrato con ese número.", "warning")
        return render_template("hc/configuracion/contrato_form.html",
                               modo="crear", contrato=form, cliente=cliente,
                               sedes=repo_sedes.listar(),
                               manuales_tarifarios=repo_manuales.listar_activos())
    try:
        repo_contratos.crear(data)
        flash("Contrato creado correctamente.", "success")
        return redirect(url_for("hc_configuracion.cliente_editar", cliente_id=cliente_id))
    except Exception as e:
        flash(f"Error al guardar el contrato: {e}", "danger")
        return render_template("hc/configuracion/contrato_form.html",
                               modo="crear", contrato=form, cliente=cliente,
                               sedes=repo_sedes.listar(),
                               manuales_tarifarios=repo_manuales.listar_activos())


@bp_hc_configuracion.route("/contratos/editar/<int:contrato_id>", methods=["GET", "POST"])
def contrato_editar(contrato_id):
    contrato = repo_contratos.obtener(contrato_id)
    if not contrato:
        flash("El contrato no existe.", "error")
        return redirect(url_for("hc_configuracion.clientes"))
    cliente = repo_clientes.obtener(contrato["cliente_id"])
    if request.method == "GET":
        return render_template("hc/configuracion/contrato_form.html",
                               modo="editar", contrato=contrato, cliente=cliente,
                               sedes=repo_sedes.listar(),
                               manuales_tarifarios=repo_manuales.listar_activos())
    form = request.form
    data = _contrato_payload(form, contrato["cliente_id"])
    if not data["nro_contrato"]:
        flash("El número de contrato es obligatorio.", "warning")
        return render_template("hc/configuracion/contrato_form.html",
                               modo="editar", contrato={**contrato, **form}, cliente=cliente,
                               sedes=repo_sedes.listar(),
                               manuales_tarifarios=repo_manuales.listar_activos())
    if repo_contratos.existe_nro(data["nro_contrato"], exclude_id=contrato_id):
        flash("Ya existe otro contrato con ese número.", "warning")
        return render_template("hc/configuracion/contrato_form.html",
                               modo="editar", contrato={**contrato, **form}, cliente=cliente,
                               sedes=repo_sedes.listar(),
                               manuales_tarifarios=repo_manuales.listar_activos())
    try:
        repo_contratos.actualizar(contrato_id, data)
        flash("Contrato actualizado correctamente.", "success")
        return redirect(url_for("hc_configuracion.contrato_editar", contrato_id=contrato_id))
    except Exception as e:
        flash(f"Error al actualizar el contrato: {e}", "danger")
        return render_template("hc/configuracion/contrato_form.html",
                               modo="editar", contrato={**contrato, **form}, cliente=cliente,
                               sedes=repo_sedes.listar(),
                               manuales_tarifarios=repo_manuales.listar_activos())


@bp_hc_configuracion.route("/contratos/toggle/<int:contrato_id>", methods=["POST"])
def contrato_toggle(contrato_id):
    contrato = repo_contratos.obtener(contrato_id)
    if not contrato:
        flash("El contrato no existe.", "error")
        return redirect(url_for("hc_configuracion.clientes"))
    nuevo = "INACTIVO" if contrato.get("estado") == "ACTIVO" else "ACTIVO"
    repo_contratos.cambiar_estado(contrato_id, nuevo)
    flash(f"Contrato {'inactivado' if nuevo == 'INACTIVO' else 'activado'} correctamente.", "success")
    return redirect(url_for("hc_configuracion.cliente_editar", cliente_id=contrato["cliente_id"]))


def _contrato_payload(form, cliente_id):
    def num(key):
        try:
            v = (form.get(key) or "").strip()
            return float(v) if v else 0.0
        except ValueError:
            return 0.0
    def fecha(key):
        v = (form.get(key) or "").strip()
        return v if v else None
    def txt(key):
        v = (form.get(key) or "").strip()
        return v if v else None
    def entero(key):
        v = (form.get(key) or "").strip()
        try:
            return int(v) if v else None
        except ValueError:
            return None
    return {
        "cliente_id":                  cliente_id,
        "nro_contrato":                txt("nro_contrato"),
        "estado":                      form.get("estado") or "ACTIVO",
        "fecha_contrato":              fecha("fecha_contrato"),
        "fec_desde":                   fecha("fec_desde"),
        "fec_hasta":                   fecha("fec_hasta"),
        "nro_referencia":              txt("nro_referencia"),
        "descripcion":                 txt("descripcion"),
        "tipo_contrato":               txt("tipo_contrato"),
        "tipo_factura":                txt("tipo_factura"),
        "prefijo_fact":                txt("prefijo_fact"),
        "periodicidad_facturacion":    txt("periodicidad_facturacion"),
        "sede_id":                     entero("sede_id"),
        "manual_tarifario":            txt("manual_tarifario"),
        "valor_contrato":              num("valor_contrato"),
        "valor_actual":                num("valor_actual"),
        "saldo":                       num("saldo"),
        "pct_descto":                  num("pct_descto"),
        "valor_ejecutado_calculado":   num("valor_ejecutado_calculado"),
        "valor_ejecutado_facturado":   num("valor_ejecutado_facturado"),
        "valor_ejecutado_calc_citas":  num("valor_ejecutado_calc_citas"),
    }


# ══════════════════════════════════════════════════════════════════
#  MANUALES TARIFARIOS
#  REGLA CRÍTICA: rutas string fijas ANTES que dinámicas <int:>
# ══════════════════════════════════════════════════════════════════

@bp_hc_configuracion.route("/manuales-tarifarios")
def manuales_tarifarios():
    data = repo_manuales.listar()
    return render_template("hc/configuracion/manuales_tarifarios.html", data=data)


@bp_hc_configuracion.route("/manuales-tarifarios/nuevo", methods=["GET", "POST"])
def manual_tarifario_nuevo():
    manuales_base = repo_manuales.listar_activos()
    if request.method == "GET":
        return render_template("hc/configuracion/manual_tarifario_form.html",
                               modo="crear", manual={}, manuales_base=manuales_base)
    form = request.form
    data = _manual_payload(form)
    if not data["codigo"]:
        flash("El código es obligatorio.", "warning")
        return render_template("hc/configuracion/manual_tarifario_form.html",
                               modo="crear", manual=form, manuales_base=manuales_base)
    if not data["nombre"]:
        flash("El nombre es obligatorio.", "warning")
        return render_template("hc/configuracion/manual_tarifario_form.html",
                               modo="crear", manual=form, manuales_base=manuales_base)
    if repo_manuales.existe_codigo(data["codigo"]):
        flash("Ya existe un manual con ese código.", "warning")
        return render_template("hc/configuracion/manual_tarifario_form.html",
                               modo="crear", manual=form, manuales_base=manuales_base)
    try:
        resultado = repo_manuales.crear(data)
        nuevo_id  = resultado[0]["id"] if resultado else None

        # ── Heredar procedimientos del manual base si se seleccionó uno ──
        if nuevo_id and data.get("manual_base_id") and data.get("pct_base"):
            try:
                insertados = repo_manuales.heredar_procedimientos_de_base(
                    manual_id      = nuevo_id,
                    manual_base_id = data["manual_base_id"],
                    pct_base       = data["pct_base"],
                )
                if insertados:
                    flash(
                        f"Manual creado. Se heredaron {insertados} procedimientos "
                        f"del manual base con {data['pct_base']}%.",
                        "success"
                    )
                else:
                    flash("Manual creado. El manual base no tiene procedimientos o ya los tenías todos.", "success")
            except Exception as ex:
                flash(f"Manual creado, pero ocurrió un error al heredar procedimientos: {ex}", "warning")
        else:
            flash("Manual tarifario creado correctamente.", "success")

        return redirect(url_for("hc_configuracion.manual_tarifario_editar", manual_id=nuevo_id))
    except Exception as e:
        flash(f"Error al guardar: {e}", "danger")
        return render_template("hc/configuracion/manual_tarifario_form.html",
                               modo="crear", manual=form, manuales_base=manuales_base)


@bp_hc_configuracion.route("/manuales-tarifarios/conteos")
def manuales_conteos():
    ids_raw = request.args.get("ids", "")
    if not ids_raw:
        return jsonify([])
    try:
        ids = [int(i) for i in ids_raw.split(",") if i.strip().isdigit()]
    except ValueError:
        return jsonify([])
    if not ids:
        return jsonify([])
    try:
        resultado = []
        sb = get_supabase_public()
        for mid in ids:
            res_p = sb.table("hc_mt_procedimientos").select("id", count="exact").eq("manual_id", mid).execute()
            res_i = sb.table("hc_mt_items").select("id", count="exact").eq("manual_id", mid).execute()
            resultado.append({"manual_id": mid, "procs": res_p.count or 0, "items": res_i.count or 0})
        return jsonify(resultado)
    except Exception as e:
        print("ERROR manuales_conteos:", e)
        return jsonify([])


@bp_hc_configuracion.route("/manuales-tarifarios/ver/<int:manual_id>")
def manual_tarifario_ver(manual_id):
    manual = repo_manuales.obtener(manual_id)
    if not manual:
        flash("Manual no encontrado.", "error")
        return redirect(url_for("hc_configuracion.manuales_tarifarios"))
    return render_template("hc/configuracion/manual_tarifario_detalle.html", manual=manual)


@bp_hc_configuracion.route("/manuales-tarifarios/editar/<int:manual_id>", methods=["GET", "POST"])
def manual_tarifario_editar(manual_id):
    manual = repo_manuales.obtener(manual_id)
    if not manual:
        flash("Manual no encontrado.", "error")
        return redirect(url_for("hc_configuracion.manuales_tarifarios"))
    manuales_base = [m for m in repo_manuales.listar_activos() if m["id"] != manual_id]
    if request.method == "GET":
        return render_template("hc/configuracion/manual_tarifario_form.html",
                               modo="editar", manual=manual, manuales_base=manuales_base)
    form = request.form
    data = _manual_payload(form)
    if not data["codigo"]:
        flash("El código es obligatorio.", "warning")
        return render_template("hc/configuracion/manual_tarifario_form.html",
                               modo="editar", manual={**manual, **form}, manuales_base=manuales_base)
    if not data["nombre"]:
        flash("El nombre es obligatorio.", "warning")
        return render_template("hc/configuracion/manual_tarifario_form.html",
                               modo="editar", manual={**manual, **form}, manuales_base=manuales_base)
    if repo_manuales.existe_codigo(data["codigo"], exclude_id=manual_id):
        flash("Ya existe otro manual con ese código.", "warning")
        return render_template("hc/configuracion/manual_tarifario_form.html",
                               modo="editar", manual={**manual, **form}, manuales_base=manuales_base)
    try:
        repo_manuales.actualizar(manual_id, data)

        # ── Re-heredar si cambió manual base o porcentaje ──
        # Solo hereda los que NO existen aún (cod_proc propio gana)
        if data.get("manual_base_id") and data.get("pct_base"):
            try:
                insertados = repo_manuales.heredar_procedimientos_de_base(
                    manual_id      = manual_id,
                    manual_base_id = data["manual_base_id"],
                    pct_base       = data["pct_base"],
                )
                if insertados:
                    flash(
                        f"Manual actualizado. Se incorporaron {insertados} procedimientos "
                        f"nuevos del manual base con {data['pct_base']}%.",
                        "success"
                    )
                else:
                    flash("Manual actualizado correctamente.", "success")
            except Exception as ex:
                flash(f"Manual actualizado, pero ocurrió un error al sincronizar procedimientos: {ex}", "warning")
        else:
            flash("Manual actualizado correctamente.", "success")

        return redirect(url_for("hc_configuracion.manual_tarifario_editar", manual_id=manual_id))
    except Exception as e:
        flash(f"Error al actualizar: {e}", "danger")
        return render_template("hc/configuracion/manual_tarifario_form.html",
                               modo="editar", manual={**manual, **form}, manuales_base=manuales_base)


@bp_hc_configuracion.route("/manuales-tarifarios/toggle/<int:manual_id>", methods=["POST"])
def manual_tarifario_toggle(manual_id):
    manual = repo_manuales.obtener(manual_id)
    if not manual:
        flash("Manual no encontrado.", "error")
        return redirect(url_for("hc_configuracion.manuales_tarifarios"))
    nuevo = "INACTIVO" if manual.get("estado") == "ACTIVO" else "ACTIVO"
    repo_manuales.cambiar_estado(manual_id, nuevo)
    flash(f"Manual {'inactivado' if nuevo == 'INACTIVO' else 'activado'} correctamente.", "success")
    return redirect(url_for("hc_configuracion.manuales_tarifarios"))


# ══════════════════════════════════════════════════════════════════
#  PROCEDIMIENTOS (AJAX) — ⚠️ string fijos ANTES que <int:manual_id>
# ══════════════════════════════════════════════════════════════════

@bp_hc_configuracion.route("/manuales-tarifarios/procedimientos/<int:proc_id>/actualizar",
                           methods=["POST"])
def mt_procedimientos_actualizar(proc_id):
    """
    Acepta todos los campos del procedimiento:
    - identificación: cod_proc, cod_cups, nombre_procedimiento, cod_factura, grupo
    - valores: valor_paquete, valor_procedimiento, valor_suministro
    - RIPS: via_ingreso, ambito_atencion, finalidad
    """
    body = request.get_json()
    try:
        def txt(k):
            v = (body.get(k) or "")
            return str(v).strip() or None

        data = {
            # Identificación (editables desde el modal)
            "cod_proc":             txt("cod_proc"),
            "cod_cups":             txt("cod_cups"),
            "nombre_procedimiento": txt("nombre_procedimiento"),
            "cod_factura":          txt("cod_factura"),
            "grupo":                txt("grupo"),
            # Valores financieros
            "valor_paquete":        float(body.get("valor_paquete") or 0),
            "valor_procedimiento":  float(body.get("valor_procedimiento") or 0),
            "valor_suministro":     float(body.get("valor_suministro") or 0),
            # Normativa RIPS
            "via_ingreso":          body.get("via_ingreso") or None,
            "ambito_atencion":      body.get("ambito_atencion") or None,
            "finalidad":            body.get("finalidad") or None,
        }
        repo_manuales.actualizar_procedimiento(proc_id, data)
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "msg": str(e)}), 400


@bp_hc_configuracion.route("/manuales-tarifarios/procedimientos/<int:proc_id>/eliminar",
                           methods=["POST"])
def mt_procedimientos_eliminar(proc_id):
    repo_manuales.eliminar_procedimiento(proc_id)
    return jsonify({"ok": True})


@bp_hc_configuracion.route("/manuales-tarifarios/<int:manual_id>/procedimientos", methods=["GET"])
def mt_procedimientos_listar(manual_id):
    return jsonify(repo_manuales.listar_procedimientos(manual_id))


@bp_hc_configuracion.route("/manuales-tarifarios/<int:manual_id>/procedimientos/agregar",
                           methods=["POST"])
def mt_procedimientos_agregar(manual_id):
    body = request.get_json()
    try:
        data = {
            "manual_id":             manual_id,
            "cups_id":               body.get("cups_id"),
            "cod_proc":              (body.get("cod_proc") or "").strip(),
            "nombre_procedimiento":  (body.get("nombre_procedimiento") or "").strip(),
            "valor_paquete":         float(body.get("valor_paquete") or 0),
            "valor_procedimiento":   float(body.get("valor_procedimiento") or 0),
            "valor_suministro":      float(body.get("valor_suministro") or 0),
            "cod_factura":           (body.get("cod_factura") or "").strip() or None,
            "cod_cups":              (body.get("cod_cups") or "").strip() or None,
            "grupo":                 (body.get("grupo") or "").strip() or None,
            "via_ingreso":           body.get("via_ingreso") or None,
            "ambito_atencion":       body.get("ambito_atencion") or None,
            "finalidad":             body.get("finalidad") or None,
        }
        repo_manuales.agregar_procedimiento(data)
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "msg": str(e)}), 400


@bp_hc_configuracion.route("/manuales-tarifarios/<int:manual_id>/procedimientos/importar",
                           methods=["POST"])
def mt_procedimientos_importar(manual_id):
    archivo = request.files.get("archivo")
    if not archivo:
        return jsonify({"ok": False, "msg": "No se recibió archivo"}), 400

    nombre    = archivo.filename.lower()
    registros = []

    def _num_co(v):
        try:
            if v is None:
                return 0.0
            s = str(v).strip().replace(" ", "")
            if "," in s and "." in s:
                s = s.replace(".", "").replace(",", ".")
            elif "," in s:
                partes = s.split(",")
                if len(partes) == 2 and len(partes[1]) <= 2:
                    s = s.replace(",", ".")
                else:
                    s = s.replace(",", "")
            return float(s) if s else 0.0
        except (ValueError, AttributeError):
            return 0.0

    def _enum_safe(v):
        if not v:
            return None
        s = str(v).strip().upper()
        return None if s in ("NA", "N/A", "NO APLICA", "-", "") else s

    if nombre.endswith(".csv"):
        try:
            raw = archivo.stream.read()
            try:
                texto = raw.decode("utf-8-sig")
            except UnicodeDecodeError:
                texto = raw.decode("latin-1")
            lineas      = texto.split("\n")
            linea_datos = lineas[1] if len(lineas) > 1 else lineas[0]
            separador   = ";" if linea_datos.count(";") > linea_datos.count(",") else ","
            header_limpio = separador.join([
                "cod_proc", "nombre_procedimiento", "valor_paquete",
                "valor_procedimiento", "valor_suministro", "cod_factura",
                "cod_cups", "grupo", "via_ingreso", "ambito_atencion", "finalidad"
            ])
            nuevo_texto = header_limpio + "\n" + "\n".join(lineas[1:])
            stream      = io.StringIO(nuevo_texto)
            reader      = csv.reader(stream, delimiter=separador, quotechar='"')
            next(reader)
            for fila in reader:
                if not fila or not any(fila):
                    continue
                f = [str(v).strip().replace("\n", " ").replace("\r", "") if v is not None else "" for v in fila]
                while len(f) < 11:
                    f.append("")
                cod = f[0]; nom = f[1]
                if not cod and not nom:
                    continue
                registros.append({
                    "cod_proc":             cod,
                    "nombre_procedimiento": nom,
                    "valor_paquete":        _num_co(f[2]),
                    "valor_procedimiento":  _num_co(f[3]),
                    "valor_suministro":     _num_co(f[4]),
                    "cod_factura":          f[5] or None,
                    "cod_cups":             f[6] or None,
                    "grupo":                f[7] or None,
                    "via_ingreso":          _enum_safe(f[8]),
                    "ambito_atencion":      _enum_safe(f[9]),
                    "finalidad":            _enum_safe(f[10]) if len(f) > 10 else None,
                })
        except Exception as exc:
            return jsonify({"ok": False, "msg": f"Error leyendo CSV: {exc}"}), 400

    elif nombre.endswith(".xlsx") or nombre.endswith(".xls"):
        try:
            import openpyxl
            wb   = openpyxl.load_workbook(archivo.stream, read_only=True, data_only=True)
            ws   = wb.active
            rows = list(ws.iter_rows(values_only=True))
        except Exception as exc:
            return jsonify({"ok": False, "msg": f"No se pudo leer el Excel: {exc}"}), 400
        if not rows:
            return jsonify({"ok": False, "msg": "Archivo vacío"}), 400
        headers = [str(h).strip() if h else "" for h in rows[0]]
        for fila in rows[1:]:
            row_dict = dict(zip(headers, fila))
            f = {}
            for k, v in row_dict.items():
                if k is None:
                    continue
                clave = str(k).strip().lower().replace(" ", "_")
                if not clave:
                    continue
                f[clave] = str(v).strip() if (v is not None and str(v).strip() != "None") else ""
            def _s(key): return f.get(key, "") or ""
            cod = _s("cod_proc"); nom = _s("nombre_procedimiento")
            if not cod and not nom:
                continue
            registros.append({
                "cod_proc":             cod,
                "nombre_procedimiento": nom,
                "valor_paquete":        _num_co(_s("valor_paquete")),
                "valor_procedimiento":  _num_co(_s("valor_procedimiento")),
                "valor_suministro":     _num_co(_s("valor_suministro")),
                "cod_factura":          _s("cod_factura") or None,
                "cod_cups":             _s("cod_cups") or None,
                "grupo":                _s("grupo") or None,
                "via_ingreso":          _enum_safe(_s("via_ingreso")),
                "ambito_atencion":      _enum_safe(_s("ambito_atencion")),
                "finalidad":            _enum_safe(_s("finalidad")),
            })
    else:
        return jsonify({"ok": False, "msg": "Formato no soportado. Use .csv o .xlsx"}), 400

    if not registros:
        return jsonify({"ok": False, "msg": "No se encontraron registros válidos en el archivo"}), 400

    try:
        total = repo_manuales.importar_procedimientos(manual_id, registros)
        return jsonify({"ok": True, "importados": total})
    except Exception as e:
        return jsonify({"ok": False, "msg": str(e)}), 500


# ══════════════════════════════════════════════════════════════════
#  ÍTEMS (AJAX) — ⚠️ string fijos ANTES que <int:manual_id>
# ══════════════════════════════════════════════════════════════════

@bp_hc_configuracion.route("/manuales-tarifarios/items/<int:item_id>/actualizar",
                           methods=["POST"])
def mt_items_actualizar(item_id):
    body = request.get_json()
    try:
        repo_manuales.actualizar_item(item_id, {
            "valor_unitario": float(body.get("valor_unitario") or 0),
        })
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "msg": str(e)}), 400


@bp_hc_configuracion.route("/manuales-tarifarios/items/<int:item_id>/eliminar",
                           methods=["POST"])
def mt_items_eliminar(item_id):
    repo_manuales.eliminar_item(item_id)
    return jsonify({"ok": True})


@bp_hc_configuracion.route("/manuales-tarifarios/<int:manual_id>/items")
def mt_items_listar(manual_id):
    return jsonify(repo_manuales.listar_items(manual_id))


@bp_hc_configuracion.route("/manuales-tarifarios/<int:manual_id>/items/agregar",
                           methods=["POST"])
def mt_items_agregar(manual_id):
    body = request.get_json()
    try:
        data = {
            "manual_id":      manual_id,
            "medicamento_id": body.get("medicamento_id"),
            "cod_item":       (body.get("cod_item") or "").strip(),
            "nombre":         (body.get("nombre") or "").strip(),
            "valor_unitario": float(body.get("valor_unitario") or 0),
        }
        repo_manuales.agregar_item(data)
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "msg": str(e)}), 400


@bp_hc_configuracion.route("/manuales-tarifarios/<int:manual_id>/items/importar",
                           methods=["POST"])
def mt_items_importar(manual_id):
    archivo = request.files.get("archivo")
    if not archivo:
        return jsonify({"ok": False, "msg": "No se recibió archivo"}), 400

    nombre    = archivo.filename.lower()
    registros = []

    if nombre.endswith(".csv"):
        try:
            raw = archivo.stream.read()
            try:
                texto = raw.decode("utf-8-sig")
            except UnicodeDecodeError:
                texto = raw.decode("latin-1")
            lineas  = texto.split("\n")
            linea_d = lineas[1] if len(lineas) > 1 else lineas[0]
            sep     = ";" if linea_d.count(";") > linea_d.count(",") else ","
            header  = sep.join(["cod_item", "nombre", "valor_unitario"])
            nuevo   = header + "\n" + "\n".join(lineas[1:])
            stream  = io.StringIO(nuevo)
            reader  = csv.reader(stream, delimiter=sep, quotechar='"')
            next(reader)
            for fila in reader:
                if not fila or not any(fila):
                    continue
                f = [str(v).strip() if v is not None else "" for v in fila]
                while len(f) < 3:
                    f.append("")
                cod = f[0]; nom = f[1]
                if not cod and not nom:
                    continue
                try:
                    val = float(f[2].replace(",", ".")) if f[2] else 0.0
                except ValueError:
                    val = 0.0
                registros.append({"cod_item": cod, "nombre": nom, "valor_unitario": val})
        except Exception as exc:
            return jsonify({"ok": False, "msg": f"Error leyendo CSV: {exc}"}), 400

    elif nombre.endswith(".xlsx") or nombre.endswith(".xls"):
        try:
            import openpyxl
            wb   = openpyxl.load_workbook(archivo.stream, read_only=True, data_only=True)
            ws   = wb.active
            rows = list(ws.iter_rows(values_only=True))
        except Exception as exc:
            return jsonify({"ok": False, "msg": f"No se pudo leer: {exc}"}), 400
        if not rows:
            return jsonify({"ok": False, "msg": "Archivo vacío"}), 400
        headers = [str(h).strip().lower().replace(" ", "_") if h else "" for h in rows[0]]
        for fila in rows[1:]:
            row = dict(zip(headers, fila))
            cod = str(row.get("cod_item") or "").strip()
            nom = str(row.get("nombre") or "").strip()
            if not cod and not nom:
                continue
            registros.append({
                "cod_item":       cod,
                "nombre":         nom,
                "valor_unitario": row.get("valor_unitario") or 0,
            })
    else:
        return jsonify({"ok": False, "msg": "Formato no soportado"}), 400

    if not registros:
        return jsonify({"ok": False, "msg": "No se encontraron registros válidos"}), 400

    try:
        total = repo_manuales.importar_items(manual_id, registros)
        return jsonify({"ok": True, "importados": total})
    except Exception as e:
        return jsonify({"ok": False, "msg": str(e)}), 500


@bp_hc_configuracion.route("/medicamentos/buscar-ajax")
def medicamentos_buscar_ajax():
    q = (request.args.get("q") or "").strip()
    if len(q) < 2:
        return jsonify([])
    try:
        res = (
            get_supabase_public()
            .table("hc_medicamentos")
            .select("id, codigo, nombre")
            .or_(f"codigo.ilike.%{q}%,nombre.ilike.%{q}%")
            .eq("estado", "ACTIVO")
            .limit(20)
            .execute()
        )
        return jsonify(res.data or [])
    except Exception:
        return jsonify([])


def _manual_payload(form):
    def txt(k):
        v = (form.get(k) or "").strip()
        return v if v else None
    def num(k, default=100.0):
        try:
            v = (form.get(k) or "").strip()
            return float(v) if v else default
        except ValueError:
            return default
    def entero(k):
        v = (form.get(k) or "").strip()
        try:
            return int(v) if v else None
        except ValueError:
            return None
    return {
        "codigo":         (form.get("codigo") or "").strip().upper(),
        "nombre":         (form.get("nombre") or "").strip(),
        "estado":          form.get("estado") or "ACTIVO",
        "manual_base_id":  entero("manual_base_id"),
        "pct_base":        num("pct_base", 100.0),
        "vigencia_desde":  txt("vigencia_desde"),
        "vigencia_hasta":  txt("vigencia_hasta"),
        "tipo_moneda":    (form.get("tipo_moneda") or "COP").strip(),
    }